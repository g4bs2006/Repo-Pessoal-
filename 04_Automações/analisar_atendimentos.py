"""
Analisa as conversas da clinica odontologica:
- Classifica se e novo lead ou paciente existente
- Identifica se houve agendamento
- Aponta motivo de nao-agendamento
- Gera relatorio Excel com cores
"""
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

PATH_XLS = r"C:\Users\gabriel.silva\Downloads\Agentes Odontológicos\04_Automações\conversas_junho_2026.xlsx"
OUTPUT   = r"C:\Users\gabriel.silva\Downloads\Agentes Odontológicos\04_Automações\analise_atendimentos_junho_v2.xlsx"

# ── Paleta de cores ─────────────────────────────────────────────────────────
COR = {
    "novo_lead":       "D6EAF8",  # azul claro
    "paciente_exist":  "D5F5E3",  # verde claro
    "indefinido":      "FDFEFE",  # branco
    "agendou":         "A9DFBF",  # verde
    "nao_agendou":     "FADBD8",  # vermelho claro
    "header":          "2C3E50",  # cinza escuro
    "sub_header":      "5D6D7E",  # cinza medio
}

# ── Palavras-chave ──────────────────────────────────────────────────────────

KW_EXISTENTE = [
    "confirmar", "confirmacao", "confirmação", "confirme", "confirmou",
    "consulta amanha", "consulta hoje", "consulta marcada",
    "cancelar", "cancelamento", "desmarcar", "remarcar", "reagendar",
    "meu tratamento", "minha consulta", "meu retorno", "minha avaliacao",
    "minha avaliação", "meu aparelho", "meu clareamento", "meu canal",
    "fazer o retorno", "voltar aqui", "recesso", "ferias", "feriado",
    "voltando dia", "retornando dia", "estarei de volta",
    "dr.", "dra.", "dr ", "dra ",
    "nao posso comparecer", "não posso comparecer",
    "nao vou poder", "não vou poder",
    "ate logo", "ate breve", "ate mais",
    "boa tarde doutor", "boa tarde doutora",
    "obrigado pelo atendimento", "obrigada pelo atendimento",
    "ja tenho consulta", "já tenho consulta",
    "meu nome e", "meu nome é",  # contexto de ja ser paciente
    "ja sou paciente", "já sou paciente",
    "continuar o tratamento", "continuacao do tratamento",
]

KW_NOVO_LEAD = [
    "quanto custa", "quanto fica", "qual o valor", "qual o preco",
    "qual o preço", "tabela de preco", "tabela de preço", "tem promocao",
    "tem promoção", "tem desconto",
    "tenho interesse", "gostaria de saber", "queria saber",
    "quero fazer", "quero agendar", "quero marcar",
    "primeira vez", "nunca fui", "nunca fiz",
    "vi no instagram", "vi no face", "vi no facebook",
    "vi no stories", "vi no anuncio", "vi no anúncio",
    "via instagram", "via facebook", "via google",
    "indicacao", "indicação", "me indicaram",
    "ola, tudo bem", "olá, tudo bem",
    "ola! tudo", "olá! tudo",
    "boa tarde, eu", "bom dia, eu",
    "quero informacoes", "quero informações",
    "clareamento dental", "implante dental", "aparelho dental",
    "faceta", "lente de contato dental",
    "invisalign", "ortodontia",
    "plano odontologico", "plano odontológico",
]

KW_AGENDOU = [
    "agendado", "agendamos", "agendei", "confirmado",
    "marcado para", "marcamos para", "marcada para",
    "dia e hora", "sua consulta sera", "sua consulta será",
    "ate la", "até lá", "a gente se ve", "a gente se vê",
    "esperamos voce", "esperamos você",
    "consulta no dia", "avaliacao no dia", "avaliação no dia",
    "fica para", "ficou para",
    "o horario ficou", "o horário ficou",
    "confirmamos para",
]

KW_NAO_RESPONDEU = [
    # sem resposta do paciente apos mensagem da clinica
]

KW_PRECO_ESTAGNADO = [
    "ainda pensando", "vou pensar", "deixa eu pensar",
    "quando eu puder", "por enquanto nao", "por enquanto não",
    "nao tenho dinheiro", "não tenho dinheiro",
    "ta caro", "tá caro", "muito caro", "acima do esperado",
    "vou ver com meu marido", "vou ver com minha esposa",
    "vou conversar em casa",
]

# ── Funcoes de analise ──────────────────────────────────────────────────────

def normalizar(texto: str) -> str:
    if not isinstance(texto, str):
        return ""
    return texto.lower().strip()


def score_keywords(texto: str, keywords: list) -> int:
    t = normalizar(texto)
    return sum(1 for kw in keywords if kw in t)


def primeira_msg_direcao(msgs: pd.DataFrame) -> str:
    if msgs.empty:
        return ""
    return msgs.iloc[0]["direcao"]


def analisar_conversa(conv_id: str, msgs: pd.DataFrame) -> dict:
    """Retorna dicionario com analise completa de uma conversa."""
    if msgs.empty or "enviado_em" not in msgs.columns:
        return {
            "tipo_contato": "Sem Mensagens", "cor_tipo": COR["indefinido"],
            "resultado": "Sem Mensagens", "cor_resultado": COR["indefinido"],
            "motivo": "Nenhuma mensagem registrada para esta conversa",
            "total_msgs": 0, "msgs_paciente": 0, "msgs_clinica": 0,
            "resumo_dialogo": "",
        }
    msgs_ord = msgs.sort_values("enviado_em").reset_index(drop=True)
    texto_total = " ".join(msgs_ord["texto"].dropna().astype(str))
    texto_norm  = normalizar(texto_total)

    n_msgs   = len(msgs_ord)
    n_inbound  = (msgs_ord["direcao"] == "TO_HUB").sum()   # paciente -> clinica
    n_outbound = (msgs_ord["direcao"] == "FROM_HUB").sum() # clinica -> paciente

    primeira_dir = primeira_msg_direcao(msgs_ord)
    # Primeira mensagem do paciente (se houver)
    primeira_do_paciente = msgs_ord[msgs_ord["direcao"] == "TO_HUB"]["texto"].iloc[0] \
        if not msgs_ord[msgs_ord["direcao"] == "TO_HUB"].empty else ""

    # ── Scores ────────────────────────────────────────────────────────────
    sc_exist = score_keywords(texto_norm, KW_EXISTENTE)
    sc_lead  = score_keywords(texto_norm, KW_NOVO_LEAD)
    sc_agend = score_keywords(texto_norm, KW_AGENDOU)
    sc_preco = score_keywords(texto_norm, KW_PRECO_ESTAGNADO)

    # Regra extra: se a PRIMEIRA mensagem veio da clinica (FROM_HUB) = confirmacao
    # quase certamente e paciente existente
    if primeira_dir == "FROM_HUB":
        sc_exist += 3

    # Recesso / aviso em massa: sem resposta do paciente
    e_recesso = "recesso" in texto_norm or "estarei de recesso" in texto_norm

    # ── Classificacao: tipo de contato ─────────────────────────────────────
    if e_recesso and n_inbound == 0:
        tipo_contato = "Comunicado (Recesso/Aviso)"
        cor_tipo     = "FFF9C4"  # amarelo claro
    elif sc_exist > sc_lead or primeira_dir == "FROM_HUB":
        tipo_contato = "Paciente Existente"
        cor_tipo     = COR["paciente_exist"]
    elif sc_lead > sc_exist:
        tipo_contato = "Novo Lead"
        cor_tipo     = COR["novo_lead"]
    else:
        tipo_contato = "Indefinido"
        cor_tipo     = COR["indefinido"]

    # ── Resultado do atendimento ─────────────────────────────────────────
    agendou = sc_agend >= 1

    # Paciente nao respondeu = clinica mandou mas paciente ficou sem resposta
    sem_resposta = n_inbound == 0 and n_outbound >= 1

    cancelou = any(kw in texto_norm for kw in [
        "cancelar", "cancelei", "cancela", "desmarcar", "desmarquei",
        "nao posso comparecer", "não posso comparecer",
        "nao vou poder", "não vou poder",
        "nao poderei comparecer", "não poderei comparecer",
    ])
    remarcou = any(kw in texto_norm for kw in [
        "remarcar", "reagendar", "remarquei", "outra data", "outra hora",
        "outra opcao", "outra opção",
    ])
    confirmou_consulta_existente = any(kw in texto_norm for kw in [
        "vou estar la", "vou estar lá", "estarei la", "estarei lá",
        "sim, confirmo", "sim confirmo", "confirmado", "ok, ate", "ok, até",
        "tudo certo", "combinado", "perfeito",
    ])

    if agendou:
        resultado       = "Agendamento Realizado"
        cor_resultado   = COR["agendou"]
    elif sem_resposta:
        resultado     = "Sem Resposta do Paciente"
        cor_resultado = "FFF9C4"
    elif e_recesso and n_inbound == 0:
        resultado     = "Comunicado sem Retorno"
        cor_resultado = "FFF9C4"
    elif cancelou and remarcou:
        resultado     = "Cancelou e Remarcou"
        cor_resultado = "F9E79F"
    elif cancelou:
        resultado     = "Cancelamento"
        cor_resultado = COR["nao_agendou"]
    elif confirmou_consulta_existente:
        resultado     = "Confirmacao de Consulta Existente"
        cor_resultado = "D5F5E3"
    elif sc_preco >= 1:
        resultado     = "Estagnado em Negociacao de Preco"
        cor_resultado = COR["nao_agendou"]
    else:
        resultado     = "Sem Agendamento"
        cor_resultado = COR["nao_agendou"]

    # ── Motivo de nao-agendamento ─────────────────────────────────────────
    if resultado == "Agendamento Realizado":
        motivo = "-"
    elif resultado == "Confirmacao de Consulta Existente":
        motivo = "Paciente ja tinha consulta marcada — apenas confirmou presença"
    elif resultado == "Cancelamento":
        motivo = "Paciente cancelou a consulta sem remarcar"
    elif resultado == "Cancelou e Remarcou":
        motivo = "Paciente cancelou mas pediu novo horario"
    elif resultado == "Sem Resposta do Paciente":
        motivo = "Clinica enviou mensagem mas paciente nao respondeu"
    elif resultado == "Comunicado sem Retorno":
        motivo = "Aviso de recesso/feriado enviado em massa, sem interacao do paciente"
    elif resultado == "Estagnado em Negociacao de Preco":
        motivo = "Lead demonstrou interesse mas nao fechou por questao de preco/tempo"
    elif tipo_contato == "Paciente Existente":
        motivo = "Paciente existente — conversa nao gerou novo agendamento"
    elif tipo_contato == "Novo Lead":
        motivo = "Lead entrou em contato mas nao houve confirmacao de agendamento"
    else:
        motivo = "Conversa encerrada sem agendamento identificado"

    # ── Resumo do dialogo ─────────────────────────────────────────────────
    # Primeiros 3 turnos do paciente para contexto
    msgs_paciente = msgs_ord[msgs_ord["direcao"] == "TO_HUB"]["texto"].dropna().tolist()
    resumo = " / ".join(str(m)[:80] for m in msgs_paciente[:3])
    if not resumo:
        msgs_clinica = msgs_ord[msgs_ord["direcao"] == "FROM_HUB"]["texto"].dropna().tolist()
        resumo = "[Apenas clinica falou] " + " / ".join(str(m)[:60] for m in msgs_clinica[:2])

    return {
        "tipo_contato":    tipo_contato,
        "cor_tipo":        cor_tipo,
        "resultado":       resultado,
        "cor_resultado":   cor_resultado,
        "motivo":          motivo,
        "total_msgs":      n_msgs,
        "msgs_paciente":   int(n_inbound),
        "msgs_clinica":    int(n_outbound),
        "resumo_dialogo":  resumo,
    }


# ── Carregamento ─────────────────────────────────────────────────────────────
print("Lendo planilha...")
df_conv = pd.read_excel(PATH_XLS, sheet_name="Conversas")
df_msgs = pd.read_excel(PATH_XLS, sheet_name="Mensagens")

# Agrupa mensagens por conversa
grupos = df_msgs.groupby("id_conversa")

print("Analisando conversas...")
rows = []
for _, conv in df_conv.iterrows():
    cid  = conv["id_conversa"]
    msgs = grupos.get_group(cid) if cid in grupos.groups else pd.DataFrame()
    analise = analisar_conversa(cid, msgs)

    # Extrai telefone
    tel = str(conv.get("contato_tel") or "")
    tel_raw = df_msgs[df_msgs["id_conversa"] == cid]["contato_tel"].dropna()
    if tel == "" and not tel_raw.empty:
        tel = str(tel_raw.iloc[0])

    rows.append({
        "ID Conversa":       cid,
        "Telefone":          tel,
        "Atendente":         conv.get("atendente", ""),
        "Status":            conv.get("status", ""),
        "Criado Em":         conv.get("criado_em", ""),
        "Encerrado Em":      conv.get("encerrado_em", ""),
        "Tipo de Contato":   analise["tipo_contato"],
        "Resultado":         analise["resultado"],
        "Motivo Nao-Agend.": analise["motivo"],
        "Total Msgs":        analise["total_msgs"],
        "Msgs Paciente":     analise["msgs_paciente"],
        "Msgs Clinica":      analise["msgs_clinica"],
        "Resumo do Dialogo": analise["resumo_dialogo"],
        # metadados internos para colorir (removidos depois)
        "_cor_tipo":         analise["cor_tipo"],
        "_cor_result":       analise["cor_resultado"],
    })

df_analise = pd.DataFrame(rows)

# ── Resumo estatistico ───────────────────────────────────────────────────────
total        = len(df_analise)
tipo_counts  = df_analise["Tipo de Contato"].value_counts()
result_counts= df_analise["Resultado"].value_counts()
motivo_counts= df_analise["Motivo Nao-Agend."].value_counts()

print(f"\nTotal conversas: {total}")
print("\nTipo de contato:")
print(tipo_counts.to_string())
print("\nResultados:")
print(result_counts.to_string())

# ── Export Excel ─────────────────────────────────────────────────────────────
print("\nGerando Excel com analise...")

# Separa metadados de cor
cor_tipo_list   = df_analise.pop("_cor_tipo").tolist()
cor_result_list = df_analise.pop("_cor_result").tolist()

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    # ── Aba 1: Analise por conversa ──────────────────────────────────────
    df_analise.to_excel(writer, sheet_name="Analise", index=False)
    ws = writer.sheets["Analise"]

    # Header
    header_fill = PatternFill("solid", fgColor=COR["header"])
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Linhas com cor condicional
    COL_TIPO   = df_analise.columns.get_loc("Tipo de Contato") + 1
    COL_RESULT = df_analise.columns.get_loc("Resultado") + 1
    COL_RESUMO = df_analise.columns.get_loc("Resumo do Dialogo") + 1

    for i, (ct, cr) in enumerate(zip(cor_tipo_list, cor_result_list), start=2):
        fill_tipo   = PatternFill("solid", fgColor=ct)
        fill_result = PatternFill("solid", fgColor=cr)
        ws.cell(i, COL_TIPO).fill   = fill_tipo
        ws.cell(i, COL_RESULT).fill = fill_result
        # Wrap text no resumo
        ws.cell(i, COL_RESUMO).alignment = Alignment(wrap_text=True, vertical="top")

    # Largura das colunas
    larguras = {
        "ID Conversa": 38, "Telefone": 20, "Atendente": 22,
        "Status": 16, "Criado Em": 22, "Encerrado Em": 22,
        "Tipo de Contato": 24, "Resultado": 30,
        "Motivo Nao-Agend.": 55, "Total Msgs": 12,
        "Msgs Paciente": 14, "Msgs Clinica": 14,
        "Resumo do Dialogo": 70,
    }
    for col_idx, col_name in enumerate(df_analise.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_name, 18)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Aba 2: Resumo Executivo ──────────────────────────────────────────
    ws2 = writer.book.create_sheet("Resumo Executivo")

    hf  = PatternFill("solid", fgColor=COR["header"])
    shf = PatternFill("solid", fgColor=COR["sub_header"])
    hft = Font(bold=True, color="FFFFFF", size=12)
    sft = Font(bold=True, color="FFFFFF", size=11)
    bft = Font(bold=True, size=11)

    def titulo(ws, row, text, span=2, fill=hf, font=hft):
        c = ws.cell(row, 1, text)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

    def dado(ws, row, label, valor, cor_val=None):
        lc = ws.cell(row, 1, label)
        lc.font = Font(bold=True)
        vc = ws.cell(row, 2, valor)
        if cor_val:
            vc.fill = PatternFill("solid", fgColor=cor_val)
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(vertical="center")

    # Cabecalho geral
    titulo(ws2, 1, "RELATORIO DE ATENDIMENTOS - JUNHO 2026 (01/06 a 19/06)", span=3)
    titulo(ws2, 2, f"Canal: (31) 99431-9431  |  Total de conversas: {total}", span=3, fill=shf, font=sft)

    r = 4
    titulo(ws2, r, "CLASSIFICACAO DE CONTATOS", span=3, fill=shf, font=sft); r+=1
    for tipo, qtd in tipo_counts.items():
        pct = round(100*qtd/total, 1)
        cor = {"Paciente Existente": COR["paciente_exist"],
               "Novo Lead": COR["novo_lead"],
               "Comunicado (Recesso/Aviso)": "FFF9C4"}.get(tipo, COR["indefinido"])
        dado(ws2, r, f"  {tipo}", f"{qtd}  ({pct}%)", cor_val=cor); r+=1

    r += 1
    titulo(ws2, r, "RESULTADOS DOS ATENDIMENTOS", span=3, fill=shf, font=sft); r+=1
    for res, qtd in result_counts.items():
        pct = round(100*qtd/total, 1)
        dado(ws2, r, f"  {res}", f"{qtd}  ({pct}%)"); r+=1

    r += 1
    titulo(ws2, r, "INTERPRETACAO E DIAGNOSTICO", span=3, fill=shf, font=sft); r+=1

    n_exist    = tipo_counts.get("Paciente Existente", 0)
    n_lead     = tipo_counts.get("Novo Lead", 0)
    n_recesso  = tipo_counts.get("Comunicado (Recesso/Aviso)", 0)
    n_agendou  = result_counts.get("Agendamento Realizado", 0)
    n_confirm  = result_counts.get("Confirmacao de Consulta Existente", 0)
    n_cancel   = result_counts.get("Cancelamento", 0)
    n_semresp  = result_counts.get("Sem Resposta do Paciente", 0)

    diagnosticos = [
        ("Volume de novos leads:",
         f"{n_lead} de {total} conversas ({round(100*n_lead/total,1)}%) sao de novos contatos. "
         f"A grande maioria ({round(100*n_exist/total,1)}%) sao pacientes ja cadastrados."),
        ("Agendamentos realizados:",
         f"{n_agendou} agendamentos confirmados no periodo. "
         f"Baixo dado o volume de conversas recebidas."),
        ("Confirmacoes de consultas:",
         f"{n_confirm} conversas sao confirmacoes de consultas ja existentes — "
         f"a IA nao precisa agendar pois ja ha consulta marcada."),
        ("Cancelamentos:",
         f"{n_cancel} cancelamentos registrados no periodo. "
         f"Avaliar taxa de cancelamento e seguimento com esses pacientes."),
        ("Sem resposta:",
         f"{n_semresp} vezes a clinica enviou mensagem e o paciente nao respondeu. "
         f"Possivel lista de pacientes inativos ou aviso de recesso."),
        ("Comunicados de recesso:",
         f"{n_recesso} mensagens foram disparos de aviso de recesso/feriado. "
         f"Nao geram agendamento por serem unidirecionais."),
        ("CAUSA RAIZ DO PROBLEMA:",
         "O canal esta recebendo PRINCIPALMENTE mensagens de pacientes EXISTENTES, "
         "nao de leads novos da META. Isso pode indicar: (1) Anuncios nao estao "
         "redirecionando para esse numero; (2) O numero foi compartilhado "
         "manualmente com pacientes antigos; (3) Ha outro canal ativo que recebe "
         "os leads da META. Recomenda-se verificar os anuncios ativos e o numero "
         "de WhatsApp Business configurado no gerenciador da META."),
    ]

    for label, texto in diagnosticos:
        lc = ws2.cell(r, 1, f"  {label}")
        lc.font = Font(bold=True, size=10)
        tc = ws2.cell(r, 2, texto)
        tc.alignment = Alignment(wrap_text=True, vertical="top")
        tc.font = Font(size=10)
        ws2.row_dimensions[r].height = 40
        if "CAUSA RAIZ" in label:
            lc.fill = PatternFill("solid", fgColor="FADBD8")
            tc.fill = PatternFill("solid", fgColor="FADBD8")
            lc.font = Font(bold=True, color="C0392B", size=10)
        r += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 90

    # ── Aba 3: Dialogo Completo ──────────────────────────────────────────
    msgs_export = df_msgs[["id_conversa","contato_tel","direcao","tipo","texto","enviado_em","status_msg"]].copy()
    msgs_export.columns = ["ID Conversa","Telefone","Direcao","Tipo","Texto","Enviado Em","Status"]
    msgs_export.to_excel(writer, sheet_name="Mensagens Completas", index=False)
    ws3 = writer.sheets["Mensagens Completas"]
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 70
    ws3.column_dimensions["F"].width = 22
    ws3.column_dimensions["G"].width = 16
    for row in ws3.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
    ws3.freeze_panes = "A2"

print(f"\nPronto! Arquivo salvo em:\n{OUTPUT}")
print(f"\nResumo:")
print(f"  Pacientes existentes : {tipo_counts.get('Paciente Existente', 0)}")
print(f"  Novos leads          : {tipo_counts.get('Novo Lead', 0)}")
print(f"  Comunicados          : {tipo_counts.get('Comunicado (Recesso/Aviso)', 0)}")
print(f"  Agendamentos         : {result_counts.get('Agendamento Realizado', 0)}")
