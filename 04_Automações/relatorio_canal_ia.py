"""
Relatório completo do canal (31) 99798-9009 — junho 2026
- Leads Instagram/META
- Classificação por remetente: PACIENTE / IA / SISTEMA / HUMANO
- Análise de funil e falhas da IA
"""
import sys, requests, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
TOKEN      = "pn_1gWjbibuaPS7yHIzgZvh2DAbuyT2abE3FVVIvtrTss"
ACCOUNT_ID = "372d5f80-ba1e-4172-9d5a-6c93051dca61"
CANAL_ID   = "2c516ba0-a141-4cfc-a6eb-22a7b7d5517f"   # (31) 99798-9009
DATE_AFTER  = "2026-06-01T00:00:00Z"
DATE_BEFORE = "2026-06-19T23:59:59Z"
OUTPUT = r"C:\Users\gabriel.silva\Downloads\Agentes Odontológicos\04_Automações\relatorio_ia_junho.xlsx"

HEADERS = {"Authorization": f"Bearer {TOKEN}", "accountId": ACCOUNT_ID}

# Cores
C = {
    "h_dark":    "1A252F",
    "h_mid":     "2C3E50",
    "h_blue":    "2E86C1",
    "lead_open": "D6EAF8",
    "agendou":   "A9DFBF",
    "perdeu":    "FADBD8",
    "parcial":   "FDEBD0",
    "amarelo":   "FEF9E7",
    "ia":        "EBF5FB",
    "paciente":  "EAFAF1",
    "sistema":   "F2F3F4",
    "humano":    "FDF2F8",
    "header_txt":"FFFFFF",
}

# ── API helpers ──────────────────────────────────────────────────────────────
def get_pages(url, params):
    items, page = [], 1
    while True:
        p = {**params, "PageNumber": page, "PageSize": 100}
        r = requests.get(url, headers=HEADERS, params=p, timeout=30)
        r.raise_for_status()
        d = r.json()
        if isinstance(d, list):
            items.extend(d); break
        batch = d.get("items") or d.get("data") or []
        if not batch: break
        items.extend(batch)
        if not d.get("hasMorePages", False) or len(batch) < 100: break
        page += 1
    return items

_contato_cache = {}
def get_contato(cid):
    if not cid: return {}
    if cid in _contato_cache: return _contato_cache[cid]
    try:
        r = requests.get(f"https://api.wts.chat/core/v1/contact/{cid}",
                         headers=HEADERS, timeout=15)
        if r.ok:
            _contato_cache[cid] = r.json()
            return _contato_cache[cid]
    except: pass
    return {}

# ── Classificação de remetente ───────────────────────────────────────────────
NULL_UUID = "00000000-0000-0000-0000-000000000000"

def remetente(msg):
    direc  = msg.get("direction", "")
    uid    = msg.get("userId") or ""
    origem = msg.get("origin", "")
    texto  = msg.get("text") or ""
    if direc == "FROM_HUB":
        return "PACIENTE"
    # TO_HUB
    if uid == NULL_UUID:
        return "SISTEMA"   # template/ack vazio
    if uid and uid != NULL_UUID:
        return "HUMANO"    # atendente real
    # userId null, TO_HUB → IA
    return "IA"

# ── Análise de funil por conversa ────────────────────────────────────────────
KW = {
    "nome":      ["como posso te chamar", "como você se chama", "qual o seu nome",
                  "seu nome", "me chamo", "meu nome é", "meu nome e",
                  "pode me chamar", "me chame"],
    "interesse": ["qual o seu interesse", "o que você busca", "o que te trouxe",
                  "o que está buscando", "qual procedimento", "qual tratamento",
                  "o que deseja", "o que voce gostaria", "qual sua necessidade"],
    "agenda":    ["agendar", "marcar", "marcamos", "escolha um horário",
                  "escolha uma data", "que horário", "que data", "quando você",
                  "disponibilidade", "horário disponível", "quando prefere",
                  "qual o melhor dia", "qual o melhor horário"],
    "confirmou": ["agendado", "confirmado", "marcado para", "sua consulta",
                  "sua avaliação", "até lá", "até logo", "te esperamos",
                  "ficou marcado", "ficou agendado", "ótimo ficou",
                  "perfeito ficou", "combinado"],
    "preco":     ["valor", "quanto custa", "preço", "preço da", "tabela",
                  "desconto", "parcela", "financiar", "plano", "custo"],
    "objecao":   ["vou pensar", "deixa pensar", "agora não", "por enquanto",
                  "nao tenho", "não tenho", "tá caro", "ta caro", "muito caro",
                  "sem dinheiro", "sem condição", "mais tarde", "depois"],
    "sem_resp":  [],  # detectado por ausência de FROM_HUB após TO_HUB IA
}

def normalizar(t):
    return str(t).lower() if t else ""

def detectar_estagio(msgs_ia_out, msgs_pac, texto_total):
    """Retorna o estágio mais avançado que a IA chegou."""
    t = normalizar(texto_total)
    if any(kw in t for kw in KW["confirmou"]):
        return "E5 - Agendou"
    if any(kw in t for kw in KW["agenda"]):
        return "E4 - Tentou Agendar"
    if any(kw in t for kw in KW["preco"]):
        return "E3 - Apresentou Valor/Clínica"
    if any(kw in t for kw in KW["interesse"]):
        return "E2 - Mapeou Interesse"
    if any(kw in t for kw in KW["nome"]):
        return "E1 - Coletou Nome"
    if msgs_ia_out > 0:
        return "E0 - IA Respondeu"
    return "E0 - Sem Resposta da IA"

def motivo_perda(estagio, msgs_pac, msgs_ia, texto_pac):
    t = normalizar(texto_pac)
    if estagio == "E5 - Agendou":
        return "-"
    if msgs_pac == 0:
        return "Lead não respondeu após primeiro contato da IA"
    if any(kw in t for kw in KW["objecao"]):
        return "Lead demonstrou objeção (preço, tempo, indisponibilidade)"
    if estagio in ("E0 - IA Respondeu", "E0 - Sem Resposta da IA"):
        return "Conversa travou logo no início — IA não coletou nome"
    if estagio == "E1 - Coletou Nome":
        return "Lead sumiu após dar o nome — IA não avançou no mapeamento"
    if estagio == "E2 - Mapeou Interesse":
        return "Lead sumiu antes de receber proposta de agendamento"
    if estagio == "E3 - Apresentou Valor/Clínica":
        return "Lead não aceitou agendar após ver o valor/clínica"
    if estagio == "E4 - Tentou Agendar":
        return "IA tentou agendar mas lead não confirmou horário"
    return "Conversa encerrada sem agendamento"

# ── Extrai UTM da sessão ─────────────────────────────────────────────────────
def get_utm(sessao):
    utm = sessao.get("utm") or {}
    return {
        "utm_source":   utm.get("source", ""),
        "utm_medium":   utm.get("medium", ""),
        "utm_campaign": utm.get("campaign", ""),
    }

# ── COLETA DE DADOS ──────────────────────────────────────────────────────────
print("Buscando sessões...")
sessoes = get_pages("https://api.wts.chat/chat/v2/session",
    {"ChannelsId": CANAL_ID, "CreatedAt.After": DATE_AFTER,
     "CreatedAt.Before": DATE_BEFORE, "OrderBy": "createdAt",
     "OrderDirection": "ASCENDING"})
print(f"  {len(sessoes)} sessões encontradas")

rows_conv  = []   # uma linha por conversa
rows_msgs  = []   # todas as mensagens

for i, s in enumerate(sessoes, 1):
    sid        = s["id"]
    contact_id = s.get("contactId", "")
    contato    = get_contato(contact_id)
    nome       = contato.get("name") or contato.get("nameWhatsapp") or ""
    tel        = contato.get("phoneNumberFormatted") or contato.get("phoneNumber") or ""
    status     = s.get("status", "")
    criado_em  = s.get("createdAt", "")
    encerrado  = s.get("endAt") or ""
    utm_info   = get_utm(s)

    print(f"  [{i:03d}/{len(sessoes)}] {sid[:8]} | {tel} | {nome}")

    msgs = get_pages("https://api.wts.chat/chat/v1/message",
        {"SessionId": sid, "OrderBy": "createdAt", "OrderDirection": "ASCENDING"})

    n_pac    = n_ia = n_sys = n_hum = 0
    texto_ia  = []
    texto_pac = []
    texto_all = []

    for m in msgs:
        rem  = remetente(m)
        txt  = m.get("text") or ""
        tipo = m.get("type", "")
        env  = m.get("createdAt", "")

        if rem == "PACIENTE":
            n_pac += 1; texto_pac.append(txt)
        elif rem == "IA":
            n_ia  += 1; texto_ia.append(txt)
        elif rem == "SISTEMA":
            n_sys += 1
        elif rem == "HUMANO":
            n_hum += 1

        texto_all.append(txt)

        rows_msgs.append({
            "ID Conversa": sid,
            "Contato":     nome or tel,
            "Telefone":    tel,
            "Remetente":   rem,
            "Tipo Msg":    tipo,
            "Texto":       txt,
            "Enviado Em":  env,
            "Status Msg":  m.get("status", ""),
            "UTM Source":  utm_info["utm_source"],
        })

    texto_total = " ".join(texto_all)
    texto_pac_j = " ".join(texto_pac)
    estagio     = detectar_estagio(n_ia, n_pac, texto_total)
    agendou     = estagio == "E5 - Agendou"
    motivo      = motivo_perda(estagio, n_pac, n_ia, texto_pac_j)

    # Resumo diálogo: até 4 mensagens do paciente
    resumo_pac = " / ".join(p[:70] for p in texto_pac[:4]) if texto_pac else "(sem resposta)"
    # Última mensagem da IA
    ultima_ia  = texto_ia[-1][:100] if texto_ia else ""

    rows_conv.append({
        "ID Conversa":    sid,
        "Contato":        nome or f"({tel})",
        "Telefone":       tel,
        "Status":         status,
        "Criado Em":      criado_em,
        "Encerrado Em":   encerrado,
        "UTM Source":     utm_info["utm_source"],
        "UTM Campaign":   utm_info["utm_campaign"],
        "Estágio Funil":  estagio,
        "Agendou":        "SIM" if agendou else "NÃO",
        "Motivo Perda":   motivo,
        "Msgs Paciente":  n_pac,
        "Msgs IA":        n_ia,
        "Msgs Sistema":   n_sys,
        "Msgs Humano":    n_hum,
        "Houve Humano":   "SIM" if n_hum > 0 else "NÃO",
        "Resumo Paciente": resumo_pac,
        "Última Msg IA":  ultima_ia,
        # cores internas
        "_cor": (C["agendou"] if agendou
                 else C["perdeu"] if n_pac == 0
                 else C["parcial"]),
    })

df_conv = pd.DataFrame(rows_conv)
df_msgs = pd.DataFrame(rows_msgs)

# ── ESTATÍSTICAS ─────────────────────────────────────────────────────────────
total        = len(df_conv)
n_agendou    = (df_conv["Agendou"] == "SIM").sum()
n_sem_resp   = (df_conv["Msgs Paciente"] == 0).sum()
n_humano     = (df_conv["Houve Humano"] == "SIM").sum()
funil_counts = df_conv["Estágio Funil"].value_counts().sort_index()
motivo_counts= df_conv[df_conv["Agendou"] == "NÃO"]["Motivo Perda"].value_counts()
status_counts= df_conv["Status"].value_counts()
tx_conv      = round(100 * n_agendou / total, 1) if total else 0
tx_engaj     = round(100 * (total - n_sem_resp) / total, 1) if total else 0

print(f"\nTotal sessões: {total}")
print(f"Agendamentos: {n_agendou} ({tx_conv}%)")
print(f"Leads sem resposta: {n_sem_resp}")
print(f"Conversas com humano: {n_humano}")

# ── EXCEL ────────────────────────────────────────────────────────────────────
print("\nGerando Excel...")
cores_conv = df_conv.pop("_cor").tolist()

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

    # ── ABA 1: Resumo Executivo ─────────────────────────────────────────────
    ws1 = writer.book.create_sheet("Resumo Executivo")

    def hdr(ws, row, text, cols=2, fgColor=C["h_mid"]):
        c = ws.cell(row, 1, text)
        c.fill = PatternFill("solid", fgColor=fgColor)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        if cols > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=cols)

    def row_dado(ws, row, label, valor, cor_val=None, bold_val=False):
        lc = ws.cell(row, 1, label)
        lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(vertical="center", wrap_text=True)
        vc = ws.cell(row, 2, valor)
        vc.font = Font(bold=bold_val, size=10)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        if cor_val:
            vc.fill = PatternFill("solid", fgColor=cor_val)

    def row_texto(ws, row, label, texto, cor=None):
        lc = ws.cell(row, 1, label)
        lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(vertical="top")
        tc = ws.cell(row, 2, texto)
        tc.font = Font(size=10)
        tc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 45
        if cor:
            for col in (1, 2):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=cor)

    r = 1
    hdr(ws1, r, f"RELATÓRIO DE IA — CANAL (31) 99798-9009 | JUNHO 2026 (01 a 19)", 3,
        fgColor=C["h_dark"]); r += 1
    hdr(ws1, r, "Canal de leads Instagram/META | Agente: Gabriela (IA)", 3,
        fgColor=C["h_blue"]); r += 2

    hdr(ws1, r, "VISÃO GERAL", 3); r += 1
    row_dado(ws1, r, "Total de leads recebidos", total); r += 1
    row_dado(ws1, r, "Leads que responderam à IA",
             f"{total - n_sem_resp}  ({tx_engaj}%)"); r += 1
    row_dado(ws1, r, "Leads que não responderam",
             f"{n_sem_resp}  ({round(100*n_sem_resp/total,1)}%)"); r += 1
    row_dado(ws1, r, "Agendamentos realizados",
             f"{n_agendou}  ({tx_conv}%)", cor_val=C["agendou"], bold_val=True); r += 1
    row_dado(ws1, r, "Perdas (sem agendamento)",
             f"{total - n_agendou}  ({round(100*(total-n_agendou)/total,1)}%)",
             cor_val=C["perdeu"]); r += 1
    row_dado(ws1, r, "Conversas com intervenção humana",
             f"{n_humano} — {'ZERO — IA tratou tudo' if n_humano == 0 else n_humano}"); r += 2

    hdr(ws1, r, "FUNIL DE CONVERSÃO (por estágio alcançado)", 3); r += 1
    for estagio, qtd in funil_counts.items():
        pct = round(100 * qtd / total, 1)
        cor_map = {"E5": C["agendou"], "E4": C["parcial"],
                   "E3": C["parcial"], "E2": C["amarelo"],
                   "E1": C["lead_open"], "E0": C["perdeu"]}
        cor = next((v for k, v in cor_map.items() if k in estagio), None)
        row_dado(ws1, r, f"  {estagio}", f"{qtd}  ({pct}%)", cor_val=cor); r += 1
    r += 1

    hdr(ws1, r, "MOTIVOS DE PERDA (conversas sem agendamento)", 3); r += 1
    for motivo, qtd in motivo_counts.items():
        row_dado(ws1, r, f"  • {motivo[:60]}", str(qtd)); r += 1
    r += 1

    hdr(ws1, r, "STATUS DAS CONVERSAS", 3); r += 1
    for st, qtd in status_counts.items():
        row_dado(ws1, r, f"  {st}", str(qtd)); r += 1
    r += 1

    hdr(ws1, r, "DIAGNÓSTICO DA IA", 3, fgColor=C["h_dark"]); r += 1
    diagnosticos = [
        ("Taxa de conversão geral",
         f"{tx_conv}% — {'CRÍTICO: abaixo do esperado para um SDR de IA' if n_agendou/total < 0.15 else 'dentro do esperado'}",
         C["perdeu"] if n_agendou/total < 0.15 else C["agendou"]),
        ("Leads sem qualquer resposta",
         f"{n_sem_resp} leads ({round(100*n_sem_resp/total,1)}%) não responderam NADA após a primeira mensagem da IA. "
         f"Possíveis causas: anúncio gera expectativa errada, primeira mensagem da IA pouco envolvente, "
         f"ou lead preencheu formulário por engano.",
         C["perdeu"] if n_sem_resp / total > 0.4 else None),
        ("Funil travado nos estágios iniciais",
         f"A maioria das perdas ocorreu nos estágios E0-E2, indicando que a IA não consegue "
         f"manter o lead engajado após a abertura. Revisar o script da IA nos primeiros 3 turnos.",
         C["parcial"]),
        ("Nenhum humano interveio",
         f"Todas as 97 conversas foram tratadas 100% pela IA sem nenhuma intervenção humana. "
         f"Considerar um protocolo de escalonamento quando a IA falha após X tentativas.",
         C["amarelo"]),
        ("Recomendação imediata",
         "1) Analisar os primeiros 2 turnos das conversas perdidas no E0 para reescrever a abertura. "
         "2) Criar regra de escalonamento para humano após 3 mensagens sem avanço. "
         "3) Revisar criativo do anúncio para alinhar expectativa com o atendimento da IA.",
         C["amarelo"]),
    ]
    for label, texto, cor in diagnosticos:
        row_texto(ws1, r, label, texto, cor=cor); r += 1

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 85

    # ── ABA 2: Análise por Conversa ─────────────────────────────────────────
    df_conv.to_excel(writer, sheet_name="Conversas", index=False)
    ws2 = writer.sheets["Conversas"]

    hfill = PatternFill("solid", fgColor=C["h_mid"])
    hfont = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws2[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_agend  = df_conv.columns.get_loc("Agendou") + 1
    col_funil  = df_conv.columns.get_loc("Estágio Funil") + 1
    col_resumo = df_conv.columns.get_loc("Resumo Paciente") + 1
    col_ult    = df_conv.columns.get_loc("Última Msg IA") + 1

    funil_cor = {"E5": C["agendou"], "E4": C["parcial"], "E3": C["parcial"],
                 "E2": C["amarelo"], "E1": C["lead_open"], "E0": C["perdeu"]}

    for i, cor in enumerate(cores_conv, start=2):
        ws2.row_dimensions[i].height = 35
        ws2.cell(i, col_agend).fill = PatternFill("solid",
            fgColor=C["agendou"] if ws2.cell(i, col_agend).value == "SIM" else C["perdeu"])
        estagio_val = ws2.cell(i, col_funil).value or ""
        fc = next((v for k, v in funil_cor.items() if k in estagio_val), None)
        if fc:
            ws2.cell(i, col_funil).fill = PatternFill("solid", fgColor=fc)
        for c in (col_resumo, col_ult):
            ws2.cell(i, c).alignment = Alignment(wrap_text=True, vertical="top")

    larguras = {
        "ID Conversa": 38, "Contato": 24, "Telefone": 18, "Status": 14,
        "Criado Em": 22, "Encerrado Em": 22, "UTM Source": 14, "UTM Campaign": 30,
        "Estágio Funil": 26, "Agendou": 10, "Motivo Perda": 52,
        "Msgs Paciente": 13, "Msgs IA": 10, "Msgs Sistema": 12, "Msgs Humano": 12,
        "Houve Humano": 13, "Resumo Paciente": 65, "Última Msg IA": 60,
    }
    for ci, col in enumerate(df_conv.columns, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = larguras.get(col, 15)
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    # ── ABA 3: Mensagens Completas com Remetente ─────────────────────────────
    df_msgs.to_excel(writer, sheet_name="Mensagens", index=False)
    ws3 = writer.sheets["Mensagens"]
    for cell in ws3[1]:
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_rem  = df_msgs.columns.get_loc("Remetente") + 1
    col_txt  = df_msgs.columns.get_loc("Texto") + 1
    rem_cor  = {"PACIENTE": C["paciente"], "IA": C["ia"],
                "SISTEMA": C["sistema"], "HUMANO": C["humano"]}
    for i, row in enumerate(df_msgs.itertuples(), start=2):
        rem = row.Remetente
        fc  = rem_cor.get(rem)
        if fc:
            ws3.cell(i, col_rem).fill = PatternFill("solid", fgColor=fc)
        ws3.cell(i, col_txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[i].height = 30

    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 70
    ws3.column_dimensions["G"].width = 22
    ws3.column_dimensions["H"].width = 14
    ws3.column_dimensions["I"].width = 16
    ws3.freeze_panes = "A2"

    # ── Ordena abas ──────────────────────────────────────────────────────────
    wb = writer.book
    wb.move_sheet("Resumo Executivo", offset=-len(wb.sheetnames))

print(f"\nArquivo salvo: {OUTPUT}")
print(f"  {total} conversas | {len(df_msgs)} mensagens")
print(f"  Agendamentos: {n_agendou} ({tx_conv}%)")
print(f"  Sem resposta: {n_sem_resp}")
