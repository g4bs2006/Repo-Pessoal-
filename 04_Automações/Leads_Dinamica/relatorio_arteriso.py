#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatório de Comportamento da IA — Conta Arte Riso (Helena / WTS.chat)

Extrai TODAS as conversas da conta, classifica o comportamento da inteligência
artificial vinculada (chatbot "Chat Arte Riso", origin=BOT), detecta
transferências/handoff e aproxima o "log de habilidades" por marcação semântica,
e gera a planilha Relatorio_ArteRiso.xlsx com 3 abas (Conversas, Timeline, Resumo).

Somente-leitura na API. Respeita o rate limit (1000 req / 2 min) com throttle e
backoff em 429. Salva um cache JSON do raw extraído para reprocessar sem rebaixar
a API.
"""

import json
import re
import time
import sys
from collections import Counter
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIGURAÇÃO ───────────────────────────────────────────────────────────────
TOKEN      = "pn_UketWXVP7W886xlDNicGbxccRnENtua28OAKTYcn3Tk"
COMPANY_ID = "4fb5c394-c53c-485f-a5b6-2eed5b390f5e"

BASE_URL = "https://api.wts.chat"
HEADERS  = {"Authorization": f"Bearer {TOKEN}"}

AQUI       = Path(__file__).parent
CACHE_PATH = AQUI / "arteriso_raw.json"
XLSX_PATH  = AQUI / "Relatorio_ArteRiso.xlsx"

DELAY_REQ = 0.18    # ~5-6 req/s, folgado para o limite de 1000/2min
TIMEOUT   = 30

# ── CLIENTE HTTP (com throttle + backoff 429) ───────────────────────────────────
def _get(path, params=None, tentativas=5):
    for i in range(tentativas):
        r = requests.get(f"{BASE_URL}{path}", headers=HEADERS, params=params, timeout=TIMEOUT)
        if r.status_code == 429:
            espera = 2 ** i
            print(f"  [429] rate limit — aguardando {espera}s...")
            time.sleep(espera)
            continue
        r.raise_for_status()
        time.sleep(DELAY_REQ)
        return r.json()
    r.raise_for_status()


def _paginar(path, params=None, page_size=100):
    """Itera todas as páginas de um endpoint paginado da Helena."""
    params = dict(params or {})
    params["PageSize"] = page_size
    page = 1
    while True:
        params["PageNumber"] = page
        data = _get(path, params)
        items = data.get("items", []) if isinstance(data, dict) else (data or [])
        for it in items:
            yield it
        if len(items) < page_size:
            break
        page += 1


# ── EXTRAÇÃO ─────────────────────────────────────────────────────────────────
def mapa_equipes():
    data = _get("/core/v1/department", {"PageSize": 100})
    items = data.get("items", data) if isinstance(data, dict) else data
    return {d["id"]: d.get("name") for d in (items or [])}


def extrair_tudo():
    print("→ Mapeando equipes...")
    deps = mapa_equipes()
    print(f"  equipes: {deps}")

    print("→ Listando sessões...")
    sessoes = list(_paginar("/chat/v2/session", {
        "OrderBy": "LastInteractionAt", "OrderDirection": "DESCENDING",
    }))
    print(f"  {len(sessoes)} sessões")

    print("→ Baixando mensagens de cada sessão...")
    total = len(sessoes)
    for i, s in enumerate(sessoes, 1):
        sid = s["id"]
        msgs = list(_paginar(f"/chat/v1/session/{sid}/message", {
            "OrderBy": "CreatedAt", "OrderDirection": "ASCENDING",
        }))
        s["_messages"] = msgs
        if i % 25 == 0 or i == total:
            print(f"  {i}/{total} sessões ({sum(len(x.get('_messages',[])) for x in sessoes[:i])} msgs)")

    raw = {"departments": deps, "sessions": sessoes}
    CACHE_PATH.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")
    print(f"  cache salvo em {CACHE_PATH.name}")
    return raw


def carregar_ou_extrair(usar_cache=True):
    if usar_cache and CACHE_PATH.exists():
        print(f"→ Usando cache {CACHE_PATH.name} (rode com --refresh para rebaixar)")
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    return extrair_tudo()


# ── CLASSIFICAÇÃO ────────────────────────────────────────────────────────────
USER_ZERO = "00000000-0000-0000-0000-000000000000"


def _eh_humano(m):
    """Operador humano = mensagem de saída com userId real OU senderId em DEFAULT."""
    u = m.get("userId")
    if u and u != USER_ZERO:
        return True
    if m.get("origin") == "DEFAULT" and m.get("senderId"):
        return True
    return False


def ator_da_msg(m):
    """Contato | Humano | IA | Sistema.

    - GATEWAY/FROM_HUB        = mensagem recebida do contato
    - userId real / senderId  = operador humano respondeu
    - origin BOT              = IA (fluxo nativo do chatbot: boas-vindas, fallback)
    - origin DEFAULT (s/ user)= IA (assistente GPT respondendo via API)
    - GATEWAY/TO_HUB, TRACK   = saída/sistema
    """
    origin = m.get("origin")
    direction = m.get("direction")
    if m.get("type") == "TRACK":
        return "Sistema"
    if direction == "FROM_HUB":
        return "Contato"
    if _eh_humano(m):
        return "Humano"
    if origin in ("BOT", "DEFAULT"):
        return "IA"
    return "Sistema"


def _coerce(v):
    """transcription/text podem vir como dict {'text': ...}; normaliza para str."""
    if isinstance(v, dict):
        return v.get("text") or v.get("value") or ""
    return v or ""


def texto_msg(m):
    t = _coerce(m.get("text"))
    if not t:
        det = m.get("details") or {}
        # áudio transcrito (transcription pode ser str ou dict {'text': ...})
        t = _coerce(det.get("transcription")) if isinstance(det, dict) else ""
        if not t and m.get("type") == "TRACK":
            t = "[início da conversa / rastreio de campanha]"
        elif not t and m.get("type") in ("AUDIO", "IMAGE", "FILE", "VIDEO"):
            t = f"[{m.get('type','').lower()}]"
    return str(t).strip()


# Marcação semântica das FALAS DA IA (aproximação do "log de habilidades")
HABILIDADES = {
    "Agendamento":      r"\b(agend|marcar|hor[áa]rio|dispon[íi]vel|confirma(r|ção)|que dia|qual dia|melhor dia|encaix)\w*",
    "Coleta de dados":  r"\b(seu nome|qual.*nome|telefone|whatsapp|nascimento|cpf|plano|conv[êe]nio|iapep|carteirinha)\w*",
    # escalonamento REAL (exclui o boilerplate "uma especialista da equipe irá assumir")
    "Pedir humano":     r"\b(atendente|secret[áa]ri|vou (te )?transferir|transfer[íi]-l|encaminhar (voc|seu)|vou chamar|chamar (um|uma) (atendente|respons))\w*",
    "Fallback":         r"\b(n[ãa]o entendi|n[ãa]o compreend|reformul|desculp|n[ãa]o consegui entender|pode repetir)\w*",
    "Saudação":         r"\b(ol[áa]|oi|bom dia|boa tarde|boa noite|seja bem|tudo bem)\w*",
    "Info procedimento": r"\b(implante|canal|pr[óo]tese|limpeza|restaura|clareamento|avalia[çc][ãa]o|or[çc]amento|valor|pre[çc]o)\w*",
}
HAB_RE = {k: re.compile(v, re.IGNORECASE) for k, v in HABILIDADES.items()}


def detectar_habilidades(falas_ia):
    achadas = []
    blob = " \n ".join(falas_ia)
    for nome, rx in HAB_RE.items():
        if rx.search(blob):
            achadas.append(nome)
    return achadas


def analisar(raw):
    deps = raw["departments"]
    DEP_GERAL = "Geral"
    linhas_conversa = []
    linhas_timeline = []

    for s in raw["sessions"]:
        sid = s["id"]
        msgs = s.get("_messages", [])
        atores = [ator_da_msg(m) for m in msgs]
        cont = Counter(atores)
        falas_ia = [texto_msg(m) for m, a in zip(msgs, atores) if a == "IA" and texto_msg(m)]

        dep_nome = deps.get(s.get("departmentId"), s.get("departmentId") or "—")
        respondeu_humano = cont.get("Humano", 0) > 0      # humano realmente digitou
        atribuido = bool(s.get("userId"))                  # conversa atribuída a um usuário
        erro_ia = dep_nome == "Erros IA"
        # handoff = saiu da equipe Geral (IA) OU foi atribuída OU humano respondeu
        transferiu = dep_nome != DEP_GERAL or atribuido or respondeu_humano
        tem_humano = respondeu_humano or atribuido

        total_msgs = len(msgs)
        msgs_ia = cont.get("IA", 0)
        pct_ia = round(100 * msgs_ia / total_msgs, 1) if total_msgs else 0.0

        utm = s.get("utm") or {}
        habilidades = detectar_habilidades(falas_ia)

        # destino do handoff
        if erro_ia:
            destino = "Erros IA"
        elif dep_nome == "CRC":
            destino = "CRC (humano)"
        elif tem_humano:
            destino = "Humano (atribuído)"
        else:
            destino = "—"

        linhas_conversa.append({
            "numero": s.get("number"),
            "session_id": sid,
            "criada_em": s.get("createdAt"),
            "encerrada_em": s.get("endAt"),
            "status": s.get("statusDescription") or s.get("status"),
            "origem_lead": utm.get("source") or s.get("origin"),
            "campanha": utm.get("campaign"),
            "equipe_final": dep_nome,
            "transferiu": "Sim" if transferiu else "Não",
            "destino_handoff": destino,
            "humano_respondeu": "Sim" if respondeu_humano else "Não",
            "atribuido_usuario": "Sim" if atribuido else "Não",
            "erro_ia": "Sim" if erro_ia else "Não",
            "msgs_total": total_msgs,
            "msgs_ia": msgs_ia,
            "msgs_contato": cont.get("Contato", 0),
            "msgs_humano": cont.get("Humano", 0),
            "pct_ia": pct_ia,
            "tempo_1a_resposta": s.get("firstResponseAt"),
            "tempo_atendimento": s.get("timeService"),
            "habilidades_detectadas": ", ".join(habilidades),
            "ultimo_texto": (s.get("lastMessageText") or "").replace("\n", " ")[:200],
            "link": s.get("previewUrl"),
        })

        for m, a in zip(msgs, atores):
            txt = texto_msg(m)
            if not txt:
                continue
            linhas_timeline.append({
                "numero": s.get("number"),
                "session_id": sid,
                "horario": m.get("createdAt"),
                "ator": a,
                "tipo": m.get("type"),
                "texto": txt.replace("\n", " ")[:500],
            })

    return linhas_conversa, linhas_timeline, deps


# ── PLANILHA ─────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")


def _escrever_aba(ws, headers, rows, larguras=None):
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def gerar_planilha(conversas, timeline, deps):
    wb = openpyxl.Workbook()

    # ── Aba Conversas ──
    ws = wb.active
    ws.title = "Conversas"
    cols = [
        ("numero", "Nº Conversa", 18), ("criada_em", "Criada em", 22),
        ("encerrada_em", "Encerrada em", 22), ("status", "Status", 14),
        ("origem_lead", "Origem", 14), ("campanha", "Campanha", 20),
        ("equipe_final", "Equipe final", 14), ("transferiu", "Transferiu?", 11),
        ("destino_handoff", "Destino handoff", 18), ("humano_respondeu", "Humano resp.?", 12),
        ("atribuido_usuario", "Atribuída?", 10),
        ("erro_ia", "Erro IA?", 9), ("msgs_total", "Msgs total", 10),
        ("msgs_ia", "Msgs IA", 9), ("msgs_contato", "Msgs contato", 11),
        ("msgs_humano", "Msgs humano", 11), ("pct_ia", "% IA", 8),
        ("tempo_1a_resposta", "1ª resposta", 22), ("tempo_atendimento", "Tempo atend.", 14),
        ("habilidades_detectadas", "Habilidades detectadas", 40),
        ("ultimo_texto", "Última mensagem", 50), ("link", "Link", 30),
    ]
    _escrever_aba(ws, [c[1] for c in cols],
                  [[c.get(k) for k, _, _ in cols] for c in conversas],
                  [w for _, _, w in cols])

    # ── Aba Timeline ──
    ws2 = wb.create_sheet("Timeline")
    tcols = [("numero", "Nº Conversa", 18), ("horario", "Horário", 22),
             ("ator", "Ator", 12), ("tipo", "Tipo", 10), ("texto", "Mensagem / Evento", 90)]
    _escrever_aba(ws2, [c[1] for c in tcols],
                  [[t.get(k) for k, _, _ in tcols] for t in timeline],
                  [w for _, _, w in tcols])

    # ── Aba Resumo ──
    ws3 = wb.create_sheet("Resumo")
    total = len(conversas)
    transferidas = sum(1 for c in conversas if c["transferiu"] == "Sim")
    com_humano = sum(1 for c in conversas if c["humano_respondeu"] == "Sim")
    atribuidas = sum(1 for c in conversas if c["atribuido_usuario"] == "Sim")
    erros = sum(1 for c in conversas if c["erro_ia"] == "Sim")
    so_ia = total - transferidas
    por_origem = Counter(c["origem_lead"] or "—" for c in conversas)
    por_equipe = Counter(c["equipe_final"] for c in conversas)
    por_status = Counter(c["status"] for c in conversas)
    hab_count = Counter()
    for c in conversas:
        for h in (c["habilidades_detectadas"].split(", ") if c["habilidades_detectadas"] else []):
            hab_count[h] += 1

    def bloco(titulo, pares):
        ws3.append([titulo, ""])
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=12, color="1F4E78")
        for k, v in pares:
            ws3.append([k, v])
        ws3.append(["", ""])

    ws3.append(["RELATÓRIO DE COMPORTAMENTO DA IA — ARTE RISO", ""])
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws3.append([f"Conta {COMPANY_ID}", ""])
    ws3.append(["", ""])
    pct = lambda n: f"{round(100*n/total,1)}%" if total else "0%"
    bloco("Visão geral", [
        ("Total de conversas", total),
        ("Resolvidas só pela IA (sem handoff)", f"{so_ia}  ({pct(so_ia)})"),
        ("Com handoff / transferência", f"{transferidas}  ({pct(transferidas)})"),
        ("Humano respondeu de fato", f"{com_humano}  ({pct(com_humano)})"),
        ("Atribuídas a um usuário", f"{atribuidas}  ({pct(atribuidas)})"),
        ("Marcadas como Erro de IA", f"{erros}  ({pct(erros)})"),
    ])
    bloco("Por equipe final", sorted(por_equipe.items(), key=lambda x: -x[1]))
    bloco("Por origem do lead", sorted(por_origem.items(), key=lambda x: -x[1]))
    bloco("Por status", sorted(por_status.items(), key=lambda x: -x[1]))
    bloco("Habilidades detectadas (nº de conversas)", sorted(hab_count.items(), key=lambda x: -x[1]))
    ws3.column_dimensions["A"].width = 48
    ws3.column_dimensions["B"].width = 20

    wb.save(XLSX_PATH)
    print(f"\n✓ Planilha gerada: {XLSX_PATH}")
    print(f"  Conversas: {total} | Só IA: {so_ia} | Transferidas: {transferidas} | Erros IA: {erros}")


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    refresh = "--refresh" in sys.argv
    raw = carregar_ou_extrair(usar_cache=not refresh)
    conversas, timeline, deps = analisar(raw)
    gerar_planilha(conversas, timeline, deps)


if __name__ == "__main__":
    main()
