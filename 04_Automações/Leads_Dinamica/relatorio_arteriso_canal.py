#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatório de IA — Arte Riso (Helena / WTS.chat)
No formato do 04_Automações/relatorio_canal_ia.py: Resumo Executivo + Conversas +
Mensagens, colorido, com funil E0-E5, motivos de perda e diagnóstico.

Diferenças desta versão (reaproveitando a referência):
  - Lê as 954 conversas do cache arteriso_raw.json (rápido). Use --refresh p/ rebaixar.
  - Classificação de remetente corrigida (IA = origin BOT/DEFAULT s/ usuário;
    HUMANO só com userId real ou senderId em DEFAULT).
  - Enriquece nome/telefone com 1 varredura paginada de /core/v1/contact (766).
  - Diagnóstico dinâmico, incluindo o gap "IA disse que agendou (E5) vs cards no CRM".

Uso:
  python relatorio_arteriso_canal.py            # usa cache
  python relatorio_arteriso_canal.py --refresh  # rebaixa as conversas da API
"""
import sys, json, time
from pathlib import Path
from collections import Counter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import requests
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
TOKEN      = "pn_UketWXVP7W886xlDNicGbxccRnENtua28OAKTYcn3Tk"
COMPANY_ID = "4fb5c394-c53c-485f-a5b6-2eed5b390f5e"
PANEL_AGENDOU = "52511dbd-88e3-4a02-aa1c-241d48c6bfdf"
STEP_AGENDOU  = "6a4764d0-159a-48c1-a703-44700e529d09"

AQUI   = Path(__file__).parent
CACHE  = AQUI / "arteriso_raw.json"
CMAP   = AQUI / "contatos_map.json"
OUTPUT = str(AQUI / "Relatorio_ArteRiso_Canal.xlsx")

HEADERS = {"Authorization": f"Bearer {TOKEN}"}
BASE = "https://api.wts.chat"

# Cores (mesma paleta da referência)
C = {
    "h_dark":"1A252F","h_mid":"2C3E50","h_blue":"2E86C1",
    "lead_open":"D6EAF8","agendou":"A9DFBF","perdeu":"FADBD8","parcial":"FDEBD0",
    "amarelo":"FEF9E7","ia":"EBF5FB","paciente":"EAFAF1","sistema":"F2F3F4","humano":"FDF2F8",
}

# ── Remetente (corrigido p/ Arte Riso) ─────────────────────────────────────────
NULL_UUID = "00000000-0000-0000-0000-000000000000"
def remetente(m):
    if m.get("type") == "TRACK": return "SISTEMA"
    if m.get("direction") == "FROM_HUB": return "PACIENTE"
    uid = m.get("userId") or ""
    if (uid and uid != NULL_UUID) or (m.get("origin") == "DEFAULT" and m.get("senderId")):
        return "HUMANO"
    if m.get("origin") in ("BOT", "DEFAULT"): return "IA"
    return "SISTEMA"

def txt_msg(m):
    t = m.get("text")
    if isinstance(t, dict): t = t.get("text") or ""
    if not t:
        det = m.get("details") or {}
        tr = det.get("transcription") if isinstance(det, dict) else None
        t = (tr.get("text") if isinstance(tr, dict) else tr) or ""
    return str(t or "")

# ── Funil E0-E5 (KW da referência) ─────────────────────────────────────────────
KW = {
 "nome":["como posso te chamar","como você se chama","qual o seu nome","seu nome","me chamo","meu nome é","pode me chamar","me chame"],
 "interesse":["qual o seu interesse","o que você busca","o que te trouxe","qual procedimento","qual tratamento","sua necessidade","o que deseja","o que voce gostaria"],
 "agenda":["agendar","marcar","escolha um horário","escolha uma data","que horário","que data","disponibilidade","horário disponível","melhor dia","melhor horário","quando prefere","quando você"],
 "confirmou":["agendado","confirmado","marcado para","sua consulta","sua avaliação","te esperamos","ficou marcado","ficou agendado","combinado","te espero","até lá"],
 "preco":["valor","quanto custa","preço","tabela","parcela","plano","custo","financiar","desconto"],
 "objecao":["vou pensar","agora não","por enquanto","não tenho","tá caro","muito caro","sem condição","mais tarde","depois","sem dinheiro"],
}
def norm(t): return str(t).lower() if t else ""
def detectar_estagio(n_ia, texto_total):
    t = norm(texto_total)
    if any(k in t for k in KW["confirmou"]): return "E5 - Agendou"
    if any(k in t for k in KW["agenda"]):    return "E4 - Tentou Agendar"
    if any(k in t for k in KW["preco"]):     return "E3 - Apresentou Valor"
    if any(k in t for k in KW["interesse"]): return "E2 - Mapeou Interesse"
    if any(k in t for k in KW["nome"]):      return "E1 - Coletou Nome"
    return "E0 - IA Respondeu" if n_ia > 0 else "E0 - Sem Resposta da IA"

def motivo_perda(est, n_pac, texto_pac):
    t = norm(texto_pac)
    if est == "E5 - Agendou": return "-"
    if n_pac == 0: return "Lead não respondeu após primeiro contato da IA"
    if any(k in t for k in KW["objecao"]): return "Lead demonstrou objeção (preço, tempo, indisponibilidade)"
    if est.startswith("E0"): return "Conversa travou logo no início — IA não coletou nome"
    if est == "E1 - Coletou Nome": return "Lead sumiu após dar o nome — IA não avançou no mapeamento"
    if est == "E2 - Mapeou Interesse": return "Lead sumiu antes de receber proposta de agendamento"
    if est == "E3 - Apresentou Valor": return "Lead não aceitou agendar após ver o valor/clínica"
    if est == "E4 - Tentou Agendar": return "IA tentou agendar mas lead não confirmou horário"
    return "Conversa encerrada sem agendamento"

# ── Coleta ──────────────────────────────────────────────────────────────────
def _paginar(path, params, key_size="PageSize"):
    out, page = [], 1
    while True:
        p = {**params, key_size: 100, "PageNumber": page}
        r = requests.get(f"{BASE}{path}", headers=HEADERS, params=p, timeout=30)
        r.raise_for_status()
        d = r.json()
        items = d.get("items", []) if isinstance(d, dict) else d
        out += items
        if len(items) < 100: break
        page += 1
        time.sleep(0.15)
    return out

def baixar_conversas():
    print("Rebaixando sessões da API...")
    sess = _paginar("/chat/v2/session", {"OrderBy":"LastInteractionAt","OrderDirection":"DESCENDING"})
    print(f"  {len(sess)} sessões; baixando mensagens...")
    for i, s in enumerate(sess, 1):
        s["_messages"] = _paginar(f"/chat/v1/session/{s['id']}/message",
                                  {"OrderBy":"CreatedAt","OrderDirection":"ASCENDING"})
        if i % 50 == 0: print(f"  {i}/{len(sess)}")
    CACHE.write_text(json.dumps({"sessions":sess}, ensure_ascii=False), encoding="utf-8")
    return sess

def mapa_contatos():
    if CMAP.exists():
        return json.loads(CMAP.read_text(encoding="utf-8"))
    print("Buscando contatos (nome/telefone)...")
    contatos = _paginar("/core/v1/contact", {})
    m = {}
    for c in contatos:
        m[c["id"]] = {
            "nome": c.get("name") or "",
            "tel": c.get("phoneNumberFormatted") or c.get("phoneNumber") or c.get("phonenumberFormatted") or "",
        }
    CMAP.write_text(json.dumps(m, ensure_ascii=False), encoding="utf-8")
    print(f"  {len(m)} contatos mapeados")
    return m

def contagem_painel():
    try:
        d = requests.get(f"{BASE}/crm/v1/panel/card", headers=HEADERS,
                         params={"PanelId":PANEL_AGENDOU,"StepId":STEP_AGENDOU,"PageSize":1}, timeout=20).json()
        return d.get("totalItems") or d.get("totalCount") or 0
    except Exception:
        return None

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    refresh = "--refresh" in sys.argv
    if refresh or not CACHE.exists():
        sess = baixar_conversas()
    else:
        print(f"Usando cache {CACHE.name} (--refresh p/ rebaixar)")
        sess = json.loads(CACHE.read_text(encoding="utf-8"))["sessions"]

    cmap = mapa_contatos()
    agendou_crm = contagem_painel()

    rows_conv, rows_msgs = [], []
    for i, s in enumerate(sess, 1):
        sid = s["id"]
        info = cmap.get(s.get("contactId"), {})
        nome = info.get("nome") or ""
        tel  = info.get("tel") or ""
        msgs = s.get("_messages", [])
        utm  = s.get("utm") or {}

        n_pac = n_ia = n_sys = n_hum = 0
        texto_all, texto_ia, texto_pac = [], [], []
        for m in msgs:
            rem = remetente(m); t = txt_msg(m)
            if   rem == "PACIENTE": n_pac += 1; texto_pac.append(t)
            elif rem == "IA":       n_ia  += 1; texto_ia.append(t)
            elif rem == "SISTEMA":  n_sys += 1
            elif rem == "HUMANO":   n_hum += 1
            texto_all.append(t)
            rows_msgs.append({
                "ID Conversa": sid, "Contato": nome or tel, "Telefone": tel,
                "Remetente": rem, "Tipo Msg": m.get("type",""), "Texto": t,
                "Enviado Em": m.get("createdAt",""), "UTM Source": utm.get("source",""),
            })

        est = detectar_estagio(n_ia, " ".join(texto_all))
        agendou = est == "E5 - Agendou"
        motivo = motivo_perda(est, n_pac, " ".join(texto_pac))
        resumo_pac = " / ".join(p[:70] for p in texto_pac[:4]) if texto_pac else "(sem resposta)"
        ultima_ia  = texto_ia[-1][:100] if texto_ia else ""

        rows_conv.append({
            "ID Conversa": sid, "Contato": nome or f"({tel})", "Telefone": tel,
            "Status": s.get("statusDescription") or s.get("status",""),
            "Criado Em": s.get("createdAt",""), "Encerrado Em": s.get("endAt") or "",
            "UTM Source": utm.get("source",""), "UTM Campaign": utm.get("campaign",""),
            "Estágio Funil": est, "Agendou": "SIM" if agendou else "NÃO",
            "Motivo Perda": motivo, "Msgs Paciente": n_pac, "Msgs IA": n_ia,
            "Msgs Sistema": n_sys, "Msgs Humano": n_hum,
            "Houve Humano": "SIM" if n_hum > 0 else "NÃO",
            "Resumo Paciente": resumo_pac, "Última Msg IA": ultima_ia,
            "_cor": (C["agendou"] if agendou else C["perdeu"] if n_pac == 0 else C["parcial"]),
        })

    df_conv = pd.DataFrame(rows_conv)
    df_msgs = pd.DataFrame(rows_msgs)

    total      = len(df_conv)
    n_agendou  = (df_conv["Agendou"] == "SIM").sum()
    n_sem_resp = (df_conv["Msgs Paciente"] == 0).sum()
    n_humano   = (df_conv["Houve Humano"] == "SIM").sum()
    funil_counts  = df_conv["Estágio Funil"].value_counts().sort_index()
    motivo_counts = df_conv[df_conv["Agendou"] == "NÃO"]["Motivo Perda"].value_counts()
    status_counts = df_conv["Status"].value_counts()
    tx_conv  = round(100*n_agendou/total, 1) if total else 0
    tx_engaj = round(100*(total-n_sem_resp)/total, 1) if total else 0

    print(f"\nTotal: {total} | E5/agendou(texto): {n_agendou} ({tx_conv}%) | sem resposta: {n_sem_resp} | humano: {n_humano}")
    if agendou_crm is not None:
        print(f"Agendou REAL no painel CRM: {agendou_crm}")

    # ── EXCEL ──
    print("Gerando Excel...")
    cores_conv = df_conv.pop("_cor").tolist()
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        ws1 = writer.book.create_sheet("Resumo Executivo")

        def hdr(ws, row, text, cols=3, fgColor=C["h_mid"]):
            c = ws.cell(row, 1, text)
            c.fill = PatternFill("solid", fgColor=fgColor)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 22
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        def row_dado(ws, row, label, valor, cor_val=None, bold_val=False):
            lc = ws.cell(row, 1, label); lc.font = Font(bold=True, size=10)
            lc.alignment = Alignment(vertical="center", wrap_text=True)
            vc = ws.cell(row, 2, valor); vc.font = Font(bold=bold_val, size=10)
            vc.alignment = Alignment(vertical="center", wrap_text=True)
            if cor_val: vc.fill = PatternFill("solid", fgColor=cor_val)
        def row_texto(ws, row, label, texto, cor=None):
            lc = ws.cell(row, 1, label); lc.font = Font(bold=True, size=10)
            lc.alignment = Alignment(vertical="top")
            tc = ws.cell(row, 2, texto); tc.font = Font(size=10)
            tc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 50
            if cor:
                for col in (1, 2): ws.cell(row, col).fill = PatternFill("solid", fgColor=cor)

        r = 1
        hdr(ws1, r, "RELATÓRIO DE IA — ARTE RISO (Teresina/PI)", 3, fgColor=C["h_dark"]); r += 1
        hdr(ws1, r, f"Conta {COMPANY_ID} | Agente: Chat Arte Riso (IA)", 3, fgColor=C["h_blue"]); r += 2

        hdr(ws1, r, "VISÃO GERAL", 3); r += 1
        row_dado(ws1, r, "Total de leads recebidos", total); r += 1
        row_dado(ws1, r, "Leads que responderam à IA", f"{total-n_sem_resp}  ({tx_engaj}%)"); r += 1
        row_dado(ws1, r, "Leads que não responderam", f"{n_sem_resp}  ({round(100*n_sem_resp/total,1)}%)"); r += 1
        row_dado(ws1, r, "IA usou linguagem de agendamento (E5)", f"{n_agendou}  ({tx_conv}%)", cor_val=C["parcial"]); r += 1
        if agendou_crm is not None:
            row_dado(ws1, r, "Agendou DE FATO (cards no painel CRM)", f"{agendou_crm}  ({round(100*agendou_crm/total,1)}%)", cor_val=C["agendou"], bold_val=True); r += 1
        row_dado(ws1, r, "Conversas com intervenção humana", f"{n_humano}"); r += 2

        hdr(ws1, r, "FUNIL DE CONVERSÃO (estágio alcançado pela IA)", 3); r += 1
        cor_map = {"E5":C["agendou"],"E4":C["parcial"],"E3":C["parcial"],"E2":C["amarelo"],"E1":C["lead_open"],"E0":C["perdeu"]}
        for est, qtd in funil_counts.items():
            pct = round(100*qtd/total, 1)
            cor = next((v for k, v in cor_map.items() if k in est), None)
            row_dado(ws1, r, f"  {est}", f"{qtd}  ({pct}%)", cor_val=cor); r += 1
        r += 1

        hdr(ws1, r, "MOTIVOS DE PERDA (conversas sem agendamento)", 3); r += 1
        for motivo, qtd in motivo_counts.items():
            row_dado(ws1, r, f"  • {motivo[:60]}", str(qtd)); r += 1
        r += 1

        hdr(ws1, r, "STATUS DAS CONVERSAS", 3); r += 1
        for st, qtd in status_counts.items():
            row_dado(ws1, r, f"  {st}", str(qtd)); r += 1
        r += 1

        hdr(ws1, r, "DIAGNÓSTICO DA IA", 3, fgColor=C["h_dark"]); r += 1
        gap = (f"A IA usou linguagem de fechamento (E5) em {n_agendou} conversas, "
               f"mas o painel CRM registra apenas {agendou_crm} agendamentos reais. "
               f"O vazamento ({n_agendou-agendou_crm} conversas) está na passagem da conversa para o CRM, "
               f"não na conversa em si — revisar a automação que cria o card.") if agendou_crm is not None else \
              "Cruzar o E5 (texto) com o painel CRM para medir o agendamento real."
        diagnosticos = [
            ("Conversão real vs. discurso",
             gap, C["perdeu"]),
            ("Leads sem qualquer resposta",
             f"{n_sem_resp} leads ({round(100*n_sem_resp/total,1)}%) não responderam após a 1ª mensagem da IA. "
             f"Revisar criativo do anúncio e a abertura da IA.",
             C["perdeu"] if n_sem_resp/total > 0.3 else None),
            ("Onde o funil trava",
             f"Maior motivo de perda: '{motivo_counts.index[0] if len(motivo_counts) else '-'}' "
             f"({motivo_counts.iloc[0] if len(motivo_counts) else 0} conversas). "
             f"É o ponto com maior retorno se a IA for ajustada.",
             C["parcial"]),
            ("Intervenção humana",
             f"{n_humano} conversas tiveram humano. O restante foi 100% IA. "
             f"Avaliar protocolo de escalonamento quando a IA falha.",
             C["amarelo"]),
            ("Recomendação imediata",
             "1) Corrigir a automação de criação de card (gap E5→CRM). "
             "2) Reescrever a abertura para reduzir leads sem resposta. "
             "3) Trabalhar o maior motivo de perda acima nos primeiros turnos.",
             C["amarelo"]),
        ]
        for label, texto, cor in diagnosticos:
            row_texto(ws1, r, label, texto, cor=cor); r += 1
        ws1.column_dimensions["A"].width = 38
        ws1.column_dimensions["B"].width = 90

        # ── Aba Conversas ──
        df_conv.to_excel(writer, sheet_name="Conversas", index=False)
        ws2 = writer.sheets["Conversas"]
        hfill = PatternFill("solid", fgColor=C["h_mid"]); hfont = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws2[1]:
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_agend = df_conv.columns.get_loc("Agendou") + 1
        col_funil = df_conv.columns.get_loc("Estágio Funil") + 1
        funil_cor = {"E5":C["agendou"],"E4":C["parcial"],"E3":C["parcial"],"E2":C["amarelo"],"E1":C["lead_open"],"E0":C["perdeu"]}
        for i, cor in enumerate(cores_conv, start=2):
            ws2.row_dimensions[i].height = 32
            ws2.cell(i, col_agend).fill = PatternFill("solid", fgColor=C["agendou"] if ws2.cell(i, col_agend).value == "SIM" else C["perdeu"])
            ev = ws2.cell(i, col_funil).value or ""
            fc = next((v for k, v in funil_cor.items() if k in ev), None)
            if fc: ws2.cell(i, col_funil).fill = PatternFill("solid", fgColor=fc)
        larguras = {"ID Conversa":38,"Contato":24,"Telefone":18,"Status":14,"Criado Em":22,"Encerrado Em":22,
                    "UTM Source":14,"UTM Campaign":28,"Estágio Funil":24,"Agendou":10,"Motivo Perda":52,
                    "Msgs Paciente":13,"Msgs IA":10,"Msgs Sistema":12,"Msgs Humano":12,"Houve Humano":13,
                    "Resumo Paciente":60,"Última Msg IA":60}
        for ci, col in enumerate(df_conv.columns, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = larguras.get(col, 15)
        ws2.row_dimensions[1].height = 28; ws2.freeze_panes = "A2"

        # ── Aba Mensagens ──
        df_msgs.to_excel(writer, sheet_name="Mensagens", index=False)
        ws3 = writer.sheets["Mensagens"]
        for cell in ws3[1]:
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        col_rem = df_msgs.columns.get_loc("Remetente") + 1
        col_txt = df_msgs.columns.get_loc("Texto") + 1
        rem_cor = {"PACIENTE":C["paciente"],"IA":C["ia"],"SISTEMA":C["sistema"],"HUMANO":C["humano"]}
        for i, row in enumerate(df_msgs.itertuples(), start=2):
            fc = rem_cor.get(row.Remetente)
            if fc: ws3.cell(i, col_rem).fill = PatternFill("solid", fgColor=fc)
            ws3.cell(i, col_txt).alignment = Alignment(wrap_text=True, vertical="top")
        for col, w in zip("ABCDEFGH", [38,24,18,12,12,70,22,14]):
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = "A2"

        wb = writer.book
        wb.move_sheet("Resumo Executivo", offset=-len(wb.sheetnames))

    print(f"\nArquivo salvo: {OUTPUT}")
    print(f"  {total} conversas | {len(df_msgs)} mensagens")

if __name__ == "__main__":
    main()
