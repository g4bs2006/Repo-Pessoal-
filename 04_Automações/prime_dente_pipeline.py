"""
Prime Dente — Pipeline de análise de conversas
01-19 junho 2026 | Canal (21) 99991-5601
Separa: IA vs Atendente (por nome) | Unidade detectada via conteúdo
"""
import requests, sys, json, time, re, os
from datetime import datetime, timezone
from collections import Counter, defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

TOKEN      = "pn_4Ysp3lQlh8rfIOx2xgqnYGPfsRxULlZCqDkMm86gA"
ACCOUNT_ID = "ad711b94-9b7c-43d6-928f-208cda372a24"
HEADERS    = {"Authorization": f"Bearer {TOKEN}", "accountId": ACCOUNT_ID}
CANAL_ID   = "eda63311-5c6c-4fc2-8702-983aade007ee"

DATE_AFTER  = "2026-06-01T00:00:00Z"
DATE_BEFORE = "2026-06-19T23:59:59Z"

OUTPUT_FILE = "prime_dente_relatorio_junho.xlsx"
CKPT_FILE   = "prime_dente_checkpoint.json"

# Departamentos conhecidos
DEPT_NAMES = {
    "733c3959-d9cd-4612-9646-422efd5e8d1c": "Geral",
    "62097181-1c85-4eb4-860d-a78a43058f11": "Equipe CRC",
    "695fc85a-8c71-4ba5-a68e-557f66f25e8f": "Erros IA",
}

# Palavras-chave para detectar unidade nas conversas
# Prime Dente tem unidades em Botafogo e Méier (RJ)
UNIT_KEYWORDS = {
    "Botafogo": ["botafogo", "bota fogo", "unidade botafogo", "prime dente botafogo",
                 "rua são clemente", "rua voluntários da pátria", "humaitá"],
    "Méier":    ["méier", "meier", "unidade méier", "prime dente méier",
                 "encantado", "riachuelo", "engenho", "norte"],
}

# ────────────────────────────────────────────────
# REQUISIÇÃO COM RETRY
# ────────────────────────────────────────────────
def req_get(url, params=None, retries=5):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, params=params, timeout=30)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt == retries - 1:
                raise
            wait = 2 ** attempt
            print(f"  [retry {attempt+1}/{retries}] {e} — aguardando {wait}s")
            time.sleep(wait)

def get_pages(url, params=None):
    items, page = [], 1
    while True:
        p = {**(params or {}), "PageNumber": page, "PageSize": 100}
        d = req_get(url, p).json()
        if isinstance(d, list):
            items.extend(d)
            break
        batch = d.get("items", [])
        items.extend(batch)
        if not d.get("hasMorePages", False) or len(batch) < 100:
            break
        page += 1
    return items

# ────────────────────────────────────────────────
# 1. USUÁRIOS (para mapear ID → nome)
# ────────────────────────────────────────────────
print("Buscando usuários (agentes)...")
users = {}
try:
    u_list = get_pages("https://api.wts.chat/core/v1/user")
    for u in u_list:
        uid  = u.get("id", "")
        name = u.get("name") or u.get("displayName") or u.get("email") or uid[:8]
        users[uid] = name
    print(f"  {len(users)} usuários mapeados: {list(users.values())}")
except Exception as e:
    print(f"  Falha ao buscar usuários: {e}")

SYSTEM_USER = "00000000-0000-0000-0000-000000000000"

def classify_sender(msg):
    uid   = msg.get("userId") or ""
    direc = msg.get("direction", "")
    if direc == "FROM_HUB":
        return "PACIENTE", "—"
    if not uid or uid == "null":
        return "IA", "Agente IA"
    if uid.startswith("00000000"):
        return "SISTEMA", "Sistema"
    name = users.get(uid, uid[:8])
    return "HUMANO", name

# ────────────────────────────────────────────────
# 2. SESSÕES
# ────────────────────────────────────────────────
print(f"\nBuscando sessões jun/2026...")
sessions = get_pages("https://api.wts.chat/chat/v2/session",
    {"ChannelsId": CANAL_ID, "CreatedAt.After": DATE_AFTER, "CreatedAt.Before": DATE_BEFORE})
print(f"  {len(sessions)} sessões encontradas")

# ────────────────────────────────────────────────
# CHECKPOINT
# ────────────────────────────────────────────────
processed = {}
if os.path.exists(CKPT_FILE):
    with open(CKPT_FILE, "r", encoding="utf-8") as f:
        processed = json.load(f)
    print(f"  Retomando checkpoint: {len(processed)} sessões já processadas")

def save_ckpt():
    with open(CKPT_FILE, "w", encoding="utf-8") as f:
        json.dump(processed, f, ensure_ascii=False)

# ────────────────────────────────────────────────
# 3. DETECTAR UNIDADE
# ────────────────────────────────────────────────
def detect_unit(texts: list[str], dept_id: str, contact_name: str = "") -> str:
    combined = (" ".join(texts) + " " + contact_name).lower()
    for unit, kws in UNIT_KEYWORDS.items():
        if any(kw in combined for kw in kws):
            return unit
    return "Sem unidade definida"

# ────────────────────────────────────────────────
# 4. PROCESSAR MENSAGENS
# ────────────────────────────────────────────────
MSG_AUTO = set([
    "Olá! Tenho interesse e queria mais informações, por favor.",
    "Olá! Vi o anúncio e quero saber como posso conquistar meu sorriso fixo",
])

def analyze_session(s, idx, total):
    sid = s["id"]
    contact_name = (s.get("contactDetails") or {}).get("name") or "—"

    # Buscar telefone
    phone = "—"
    contact_id = s.get("contactId")
    if contact_id:
        try:
            cr = req_get(f"https://api.wts.chat/core/v1/contact/{contact_id}")
            cd = cr.json()
            phone = cd.get("phoneNumberFormatted") or cd.get("phone") or "—"
            if contact_name == "—":
                contact_name = cd.get("name") or "—"
        except:
            pass

    # Mensagens
    msgs = get_pages("https://api.wts.chat/chat/v1/message",
        {"SessionId": sid, "OrderBy": "createdAt", "OrderDirection": "ASCENDING"})

    texts_all = []
    rows = []
    for m in msgs:
        txt   = (m.get("text") or "").strip()
        kind, agent_name = classify_sender(m)
        dt_raw = m.get("createdAt") or m.get("sentAt") or ""
        dt_str = dt_raw[:19].replace("T", " ") if dt_raw else ""
        rows.append({
            "session_id":   sid,
            "data_hora":    dt_str,
            "remetente":    kind,
            "agente_nome":  agent_name,
            "texto":        txt,
        })
        if txt and txt not in MSG_AUTO:
            texts_all.append(txt)

    dept_id = s.get("departmentId") or ""
    unit    = detect_unit(texts_all, dept_id, contact_name)

    # Contagens
    cnt = Counter(r["remetente"] for r in rows)
    agents_in_sess = set(r["agente_nome"] for r in rows if r["remetente"] == "HUMANO")

    # Último status
    status = s.get("status") or s.get("statusDescription") or "—"

    # Agendou? (detecção por keyword)
    AGENDAR_KWS = ["agend", "marcad", "confirm", "horário definido", "te espero",
                   "dia marcado", "consulta marcada", "vou te esperar"]
    agendou = any(any(kw in r["texto"].lower() for kw in AGENDAR_KWS)
                  for r in rows if r["remetente"] in ("IA", "HUMANO"))

    print(f"  [{idx:03d}/{total}] {phone} | {contact_name} | {unit} | "
          f"IA:{cnt.get('IA',0)} H:{cnt.get('HUMANO',0)} P:{cnt.get('PACIENTE',0)}")

    return {
        "meta": {
            "session_id":    sid,
            "numero":        s.get("number"),
            "contato":       contact_name,
            "telefone":      phone,
            "status":        status,
            "departamento":  DEPT_NAMES.get(dept_id, dept_id[:8] if dept_id else "—"),
            "unidade":       unit,
            "criado_em":     s.get("createdAt", "")[:10],
            "agendou":       agendou,
            "msgs_ia":       cnt.get("IA", 0),
            "msgs_humano":   cnt.get("HUMANO", 0),
            "msgs_paciente": cnt.get("PACIENTE", 0),
            "msgs_sistema":  cnt.get("SISTEMA", 0),
            "agentes":       ", ".join(sorted(agents_in_sess)) or "—",
            "teve_humano":   len(agents_in_sess) > 0,
        },
        "msgs": rows,
    }

# ── Loop principal ──
all_sessions_data = []
for i, s in enumerate(sessions, 1):
    sid = s["id"]
    if sid in processed:
        all_sessions_data.append(processed[sid])
        continue
    result = analyze_session(s, i, len(sessions))
    processed[sid] = result
    all_sessions_data.append(result)
    if i % 20 == 0:
        save_ckpt()

save_ckpt()
print(f"\nTotal processado: {len(all_sessions_data)} sessões")

# ────────────────────────────────────────────────
# 5. MONTAR DATAFRAMES
# ────────────────────────────────────────────────
metas = [d["meta"] for d in all_sessions_data]
df    = pd.DataFrame(metas)

all_msgs = []
for d in all_sessions_data:
    unit = d["meta"]["unidade"]
    for m in d["msgs"]:
        all_msgs.append({**m, "unidade": unit, "contato": d["meta"]["contato"], "telefone": d["meta"]["telefone"]})
df_msgs = pd.DataFrame(all_msgs) if all_msgs else pd.DataFrame()

# ────────────────────────────────────────────────
# 6. STATS POR UNIDADE
# ────────────────────────────────────────────────
def stats_for(subset: pd.DataFrame) -> dict:
    if subset.empty:
        return {}
    total    = len(subset)
    agendou  = subset["agendou"].sum()
    humano   = subset["teve_humano"].sum()
    msgs_ia  = subset["msgs_ia"].sum()
    msgs_h   = subset["msgs_humano"].sum()
    msgs_p   = subset["msgs_paciente"].sum()
    msgs_s   = subset["msgs_sistema"].sum()
    total_m  = msgs_ia + msgs_h + msgs_p + msgs_s
    return {
        "total_sessoes": total,
        "agendamentos":  int(agendou),
        "tx_conv_%":     round(agendou/total*100, 1) if total else 0,
        "c_intervencao_humana": int(humano),
        "tx_humano_%":  round(humano/total*100, 1) if total else 0,
        "msgs_ia":      int(msgs_ia),
        "msgs_humano":  int(msgs_h),
        "msgs_paciente":int(msgs_p),
        "msgs_sistema": int(msgs_s),
        "total_msgs":   int(total_m),
        "pct_ia_%":     round(msgs_ia/total_m*100,1) if total_m else 0,
        "pct_humano_%": round(msgs_h/total_m*100,1) if total_m else 0,
    }

units  = df["unidade"].unique().tolist()
global_stats = stats_for(df)

print("\n=== Stats globais ===")
print(json.dumps(global_stats, ensure_ascii=False, indent=2))
for u in units:
    us = stats_for(df[df["unidade"] == u])
    print(f"\n=== {u} ===")
    print(json.dumps(us, ensure_ascii=False, indent=2))

# ────────────────────────────────────────────────
# 7. EXCEL
# ────────────────────────────────────────────────
print(f"\nGerando {OUTPUT_FILE}...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # ── Aba 1: Resumo Executivo ──────────────────
    rows_res = [["Métrica", "Global"] + units]
    metrics = [
        ("Sessões totais",            "total_sessoes"),
        ("Agendamentos",              "agendamentos"),
        ("Taxa de conversão (%)",     "tx_conv_%"),
        ("Com intervenção humana",    "c_intervencao_humana"),
        ("Taxa com humano (%)",       "tx_humano_%"),
        ("Mensagens da IA",           "msgs_ia"),
        ("Mensagens de atendentes",   "msgs_humano"),
        ("Mensagens de pacientes",    "msgs_paciente"),
        ("Mensagens do sistema",      "msgs_sistema"),
        ("Total de mensagens",        "total_msgs"),
        ("% msgs da IA",              "pct_ia_%"),
        ("% msgs de atendentes",      "pct_humano_%"),
    ]
    unit_stats_map = {u: stats_for(df[df["unidade"] == u]) for u in units}
    for label, key in metrics:
        row = [label, global_stats.get(key, "—")]
        for u in units:
            row.append(unit_stats_map[u].get(key, "—"))
        rows_res.append(row)

    df_res = pd.DataFrame(rows_res[1:], columns=rows_res[0])
    df_res.to_excel(writer, sheet_name="Resumo Executivo", index=False)

    # ── Aba 2: Análise por conversa ─────────────
    df_analise = df[[
        "unidade","departamento","criado_em","contato","telefone",
        "status","agendou","msgs_ia","msgs_humano","msgs_paciente",
        "msgs_sistema","agentes","teve_humano","session_id"
    ]].copy()
    df_analise.columns = [
        "Unidade","Departamento","Data","Contato","Telefone",
        "Status","Agendou","Msgs IA","Msgs Atendente","Msgs Paciente",
        "Msgs Sistema","Atendentes","Teve Humano","Session ID"
    ]
    df_analise.to_excel(writer, sheet_name="Conversas", index=False)

    # ── Aba 3: Intervenções humanas ──────────────
    humano_rows = df[df["teve_humano"] == True].copy()
    if not humano_rows.empty:
        df_hum = humano_rows[[
            "unidade","data_criado_em" if "data_criado_em" in humano_rows.columns else "criado_em",
            "contato","telefone","agentes","msgs_ia","msgs_humano","msgs_paciente","agendou"
        ]].copy() if "data_criado_em" in humano_rows.columns else humano_rows[[
            "unidade","criado_em","contato","telefone","agentes","msgs_ia","msgs_humano","msgs_paciente","agendou"
        ]].copy()
        df_hum.columns = [
            "Unidade","Data","Contato","Telefone","Atendente(s)",
            "Msgs IA","Msgs Atendente","Msgs Paciente","Agendou"
        ]
        df_hum.to_excel(writer, sheet_name="Intervenções Humanas", index=False)

    # ── Aba 4: Mensagens completas ───────────────
    if not df_msgs.empty:
        df_msgs_out = df_msgs[[
            "unidade","contato","telefone","data_hora","remetente","agente_nome","texto","session_id"
        ]].copy()
        df_msgs_out.columns = [
            "Unidade","Contato","Telefone","Data/Hora","Remetente","Nome Agente","Texto","Session ID"
        ]
        df_msgs_out.to_excel(writer, sheet_name="Mensagens Completas", index=False)

print("Excel base gerado.")

# ────────────────────────────────────────────────
# 8. FORMATAÇÃO
# ────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)

COLORS = {
    "header":    "1A2240",
    "ia":        "D6E9F8",   # azul claro
    "humano":    "E8D5F5",   # roxo claro
    "paciente":  "D4EDDA",   # verde claro
    "sistema":   "F5F5F5",   # cinza
    "agendou":   "C6EFCE",   # verde agendamento
    "nao_ag":    "FFDDC1",   # laranja sem agendamento
    "unit_a":    "EBF5FB",   # unidade A
    "unit_b":    "FEF9E7",   # unidade B
    "unit_c":    "F9EBEA",   # sem unidade
    "title_txt": "FFFFFF",
    "bold_txt":  "1A2240",
}

def header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.font = Font(bold=True, color=COLORS["title_txt"], size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, max_w=60):
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 2, max_w)

# ── Resumo Executivo ──
ws = wb["Resumo Executivo"]
header_row(ws)
auto_width(ws)

# ── Conversas ──
ws2 = wb["Conversas"]
header_row(ws2)
# Cor por unidade e agendamento
unit_list = df["unidade"].unique().tolist()
unit_color_map = {}
unit_colors_pool = [COLORS["unit_a"], COLORS["unit_b"], COLORS["unit_c"]]
for j, u in enumerate(unit_list):
    unit_color_map[u] = unit_colors_pool[j % len(unit_colors_pool)]

for row in ws2.iter_rows(min_row=2):
    unit_val  = row[0].value if row[0].value else ""
    agendou_v = row[6].value
    bg = unit_color_map.get(unit_val, COLORS["unit_c"])
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=bg)
    if agendou_v == True:
        row[6].fill = PatternFill("solid", fgColor=COLORS["agendou"])
        row[6].font = Font(bold=True, color="276221")
    elif agendou_v == False:
        row[6].fill = PatternFill("solid", fgColor=COLORS["nao_ag"])

    # Destacar cols de msgs
    row[7].fill  = PatternFill("solid", fgColor=COLORS["ia"])      # Msgs IA
    row[8].fill  = PatternFill("solid", fgColor=COLORS["humano"])   # Msgs Atendente
    row[9].fill  = PatternFill("solid", fgColor=COLORS["paciente"]) # Msgs Paciente

auto_width(ws2)

# ── Intervenções Humanas ──
if "Intervenções Humanas" in wb.sheetnames:
    ws3 = wb["Intervenções Humanas"]
    header_row(ws3)
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=COLORS["humano"])
        agendou_v = row[8].value if len(row) > 8 else None
        if agendou_v == True:
            row[8].fill = PatternFill("solid", fgColor=COLORS["agendou"])
            row[8].font = Font(bold=True, color="276221")
    auto_width(ws3)

# ── Mensagens Completas ──
if "Mensagens Completas" in wb.sheetnames:
    ws4 = wb["Mensagens Completas"]
    header_row(ws4)
    REM_COLORS = {
        "IA":       COLORS["ia"],
        "HUMANO":   COLORS["humano"],
        "PACIENTE": COLORS["paciente"],
        "SISTEMA":  COLORS["sistema"],
    }
    for row in ws4.iter_rows(min_row=2):
        rem = row[4].value if len(row) > 4 else ""
        bg  = REM_COLORS.get(rem, COLORS["sistema"])
        for cell in row:
            cell.fill    = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if rem == "HUMANO":
            row[5].font = Font(bold=True)  # nome do agente em negrito
    ws4.row_dimensions  # auto-height handled by wrap
    auto_width(ws4, max_w=80)

wb.save(OUTPUT_FILE)
print(f"\nArquivo salvo: {OUTPUT_FILE}")
print(f"Sessões: {len(df)} | Unidades: {df['unidade'].value_counts().to_dict()}")
print(f"Agendamentos: {df['agendou'].sum()} | Com humano: {df['teve_humano'].sum()}")
print("Concluído.")
