"""
Pipeline completo - OBClinic Odontologia - Joinville/SC
Canal (47) 99700-8423 | Agente: Gi (IA) | Junho 2026
"""
import sys, requests, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Config ───────────────────────────────────────────────────────────────────
TOKEN      = "pn_gGIh0bpd6Al6jubejrT98ToHW7aavEn4RoY8Gg"
ACCOUNT_ID = "3e35f9a6-4f6b-430c-9f80-b7b4b41fbca7"
CANAL_ID   = "223b4da1-1e1f-4403-9fd3-bfa0a83c68b6"
DATE_AFTER  = "2026-06-01T00:00:00Z"
DATE_BEFORE = "2026-06-19T23:59:59Z"
CLINICA    = "OBClinic Odontologia"
AGENTE     = "Gi"
CIDADE     = "Joinville/SC"
CANAL_NUM  = "(47) 99700-8423"

OUT_EXCEL = r"C:\Users\gabriel.silva\Downloads\Agentes Odontológicos\04_Automações\obclinic_relatorio_junho.xlsx"
OUT_JSON  = r"C:\Users\gabriel.silva\Downloads\Agentes Odontológicos\04_Automações\obclinic_stats.json"
OUT_CKPT  = r"C:\Users\gabriel.silva\Downloads\Agentes Odontológicos\04_Automações\obclinic_checkpoint.json"

HEADERS = {"Authorization": f"Bearer {TOKEN}", "accountId": ACCOUNT_ID}

# Mensagens automáticas do botão do anúncio (variantes)
MSGS_AUTO = [
    "Olá! Vi o anúncio e quero saber como posso conquistar meu sorriso fixo",
    "Olá! Vim do anúncio e quero transformar meu sorriso com lentes e facetas.",
    "Olá! Vi o anúncio e quero saber como posso Oi conquistar meu sorriso fixo",
]
def e_msg_auto(txt):
    t = str(txt or "").strip()
    return any(t.startswith(m[:40]) for m in MSGS_AUTO)

# ── API helpers ──────────────────────────────────────────────────────────────
import time

def req_get(url, params=None, timeout=30, retries=5):
    """GET com retry exponencial em erros de rede."""
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, params=params, timeout=timeout)
            r.raise_for_status()
            return r
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            wait = 2 ** attempt
            print(f"    [retry {attempt+1}/{retries}] erro de rede, aguardando {wait}s... ({e})")
            time.sleep(wait)
    raise RuntimeError(f"Falhou após {retries} tentativas: {url}")

def get_pages(url, params):
    items, page = [], 1
    while True:
        p = {**params, "PageNumber": page, "PageSize": 100}
        d = req_get(url, p).json()
        if isinstance(d, list):
            items.extend(d); break
        batch = d.get("items") or d.get("data") or []
        if not batch: break
        items.extend(batch)
        if not d.get("hasMorePages", False) or len(batch) < 100: break
        page += 1
    return items

_cache = {}
def get_contato(cid):
    if not cid: return {}
    if cid in _cache: return _cache[cid]
    try:
        r = req_get(f"https://api.wts.chat/core/v1/contact/{cid}", timeout=15)
        _cache[cid] = r.json()
        return _cache[cid]
    except: pass
    return {}

# ── Remetente ────────────────────────────────────────────────────────────────
NULL_UUID = "00000000-0000-0000-0000-000000000000"
def remetente(msg):
    direc = msg.get("direction", "")
    uid   = msg.get("userId") or ""
    if direc == "FROM_HUB": return "PACIENTE"
    if uid == NULL_UUID:    return "SISTEMA"
    if uid and uid != NULL_UUID: return "HUMANO"
    return "IA"

# ── Funil ────────────────────────────────────────────────────────────────────
KW = {
    "nome":      ["como posso te chamar","como você se chama","qual o seu nome",
                  "pode me chamar","me chame","seu nome","meu nome é","meu nome e"],
    "interesse": ["qual o seu interesse","o que você busca","o que te trouxe",
                  "o que está buscando","qual procedimento","qual tratamento",
                  "o que deseja","o que incomoda","no seu sorriso","mudar uma coisa",
                  "mudar no seu sorriso"],
    "agenda":    ["agendar","marcar","marcamos","escolha um horário","escolha uma data",
                  "que horário","que data","quando você","disponibilidade",
                  "horário disponível","quando prefere","qual o melhor dia",
                  "próximo passo","avaliação gratuita","avaliação com"],
    "confirmou": ["agendado","confirmado","marcado para","sua consulta","sua avaliação",
                  "até lá","até logo","te esperamos","ficou marcado","ficou agendado",
                  "ótimo ficou","perfeito ficou","combinado","consulta confirmada"],
    "preco":     ["valor","quanto custa","preço","tabela","desconto","parcela",
                  "financiar","plano","custo","investimento"],
    "objecao":   ["vou pensar","deixa pensar","agora não","por enquanto",
                  "nao tenho","não tenho","tá caro","ta caro","muito caro",
                  "sem dinheiro","sem condição","mais tarde","depois"],
}

def normalizar(t): return str(t).lower() if t else ""

def estagio(texto_all, n_ia):
    t = normalizar(texto_all)
    if any(k in t for k in KW["confirmou"]): return "E5 - Agendou"
    if any(k in t for k in KW["agenda"]):    return "E4 - Tentou Agendar"
    if any(k in t for k in KW["preco"]):     return "E3 - Apresentou Valor/Clínica"
    if any(k in t for k in KW["interesse"]): return "E2 - Mapeou Interesse"
    if any(k in t for k in KW["nome"]):      return "E1 - Coletou Nome"
    if n_ia > 0:                             return "E0 - IA Respondeu"
    return "E0 - Sem Resposta da IA"

def motivo(est, n_pac, texto_pac):
    t = normalizar(texto_pac)
    if est == "E5 - Agendou": return "-"
    if n_pac == 0:   return "Lead não respondeu após primeiro contato da IA"
    if any(k in t for k in KW["objecao"]): return "Lead demonstrou objeção (preço, tempo, indisponibilidade)"
    if est == "E0 - IA Respondeu": return "Conversa travou logo no início — IA não coletou nome"
    if est == "E1 - Coletou Nome": return "Lead sumiu após dar o nome — IA não avançou no mapeamento"
    if est == "E2 - Mapeou Interesse": return "Lead sumiu antes de receber proposta de agendamento"
    if est == "E3 - Apresentou Valor/Clínica": return "Lead não aceitou agendar após ver o valor/clínica"
    if est == "E4 - Tentou Agendar": return "IA tentou agendar mas lead não confirmou horário"
    return "Conversa encerrada sem agendamento"

def get_utm(s):
    utm = s.get("utm") or {}
    return utm.get("source",""), utm.get("medium",""), utm.get("campaign","")

# ── COLETA ───────────────────────────────────────────────────────────────────
print(f"Buscando sessoes do canal {CANAL_NUM}...")
sessoes = get_pages("https://api.wts.chat/chat/v2/session",
    {"ChannelsId": CANAL_ID, "CreatedAt.After": DATE_AFTER,
     "CreatedAt.Before": DATE_BEFORE, "OrderBy": "createdAt",
     "OrderDirection": "ASCENDING"})
print(f"  {len(sessoes)} sessoes encontradas")

# ── Carrega checkpoint se existir ────────────────────────────────────────────
import os
rows_conv = []
rows_msgs = []
processed_ids = set()

if os.path.exists(OUT_CKPT):
    with open(OUT_CKPT, encoding="utf-8") as f:
        ckpt = json.load(f)
    rows_conv = ckpt.get("rows_conv", [])
    rows_msgs = ckpt.get("rows_msgs", [])
    processed_ids = set(ckpt.get("processed_ids", []))
    print(f"  Retomando de checkpoint: {len(processed_ids)} sessoes ja processadas")

def salvar_checkpoint():
    with open(OUT_CKPT, "w", encoding="utf-8") as f:
        json.dump({"rows_conv": rows_conv, "rows_msgs": rows_msgs,
                   "processed_ids": list(processed_ids)}, f, ensure_ascii=False)

for i, s in enumerate(sessoes, 1):
    if s["id"] in processed_ids:
        continue
    sid    = s["id"]
    cid    = s.get("contactId","")
    cont   = get_contato(cid)
    nome   = cont.get("name") or cont.get("nameWhatsapp") or ""
    tel    = cont.get("phoneNumberFormatted") or cont.get("phoneNumber") or ""
    status = s.get("status","")
    criado = s.get("createdAt","")
    fim    = s.get("endAt","") or ""
    utm_src, utm_med, utm_camp = get_utm(s)

    if i % 50 == 0 or i == 1:
        print(f"  [{i:03d}/{len(sessoes)}] {tel} | {nome[:30]}")

    msgs = get_pages("https://api.wts.chat/chat/v1/message",
        {"SessionId": sid, "OrderBy": "createdAt", "OrderDirection": "ASCENDING"})

    n_pac = n_ia = n_sys = n_hum = 0
    txt_ia = []; txt_pac = []; txt_all = []

    for m in msgs:
        rem  = remetente(m)
        txt  = m.get("text") or ""
        tipo = m.get("type","")
        env  = m.get("createdAt","")
        if rem == "PACIENTE":  n_pac += 1; txt_pac.append(txt)
        elif rem == "IA":      n_ia  += 1; txt_ia.append(txt)
        elif rem == "SISTEMA": n_sys += 1
        elif rem == "HUMANO":  n_hum += 1
        txt_all.append(txt)
        rows_msgs.append({
            "ID Conversa": sid, "Contato": nome or tel, "Telefone": tel,
            "Remetente": rem, "Tipo Msg": tipo, "Texto": txt,
            "Enviado Em": env, "Status Msg": m.get("status",""),
            "UTM Source": utm_src,
        })

    texto_total = " ".join(txt_all)
    texto_pac_j = " ".join(txt_pac)
    # Exclui msgs automáticas do botão
    pac_reais = [t for t in txt_pac if not e_msg_auto(t)]

    est  = estagio(texto_total, n_ia)
    agnd = est == "E5 - Agendou"
    mot  = motivo(est, len(pac_reais), texto_pac_j)
    vacuo = n_ia > 0 and len(pac_reais) == 0

    resumo_pac = " / ".join(p[:70] for p in txt_pac[:4]) if txt_pac else "(sem resposta)"
    ultima_ia  = txt_ia[-1][:100] if txt_ia else ""

    processed_ids.add(sid)
    if i % 50 == 0:
        salvar_checkpoint()
        print(f"  >> checkpoint salvo ({len(processed_ids)} processadas)")

    rows_conv.append({
        "ID Conversa":    sid,
        "Contato":        nome or f"({tel})",
        "Telefone":       tel,
        "Status":         status,
        "Criado Em":      criado,
        "Encerrado Em":   fim,
        "UTM Source":     utm_src,
        "UTM Campaign":   utm_camp,
        "Estágio Funil":  est,
        "Agendou":        "SIM" if agnd else "NÃO",
        "Vacuo":          "SIM" if vacuo else "NÃO",
        "Motivo Perda":   mot,
        "Msgs Paciente":  n_pac,
        "Msgs IA":        n_ia,
        "Msgs Sistema":   n_sys,
        "Msgs Humano":    n_hum,
        "Houve Humano":   "SIM" if n_hum > 0 else "NÃO",
        "Resumo Paciente": resumo_pac,
        "Última Msg IA":  ultima_ia,
        "_cor": "#A9DFBF" if agnd else ("#FADBD8" if vacuo else "#FDEBD0"),
    })

df_conv = pd.DataFrame(rows_conv)
df_msgs = pd.DataFrame(rows_msgs)

# ── ESTATÍSTICAS ─────────────────────────────────────────────────────────────
total     = len(df_conv)
n_agend   = (df_conv["Agendou"] == "SIM").sum()
n_vacuo   = (df_conv["Vacuo"]   == "SIM").sum()
n_humano  = (df_conv["Houve Humano"] == "SIM").sum()
n_engaj   = total - n_vacuo
tx_conv   = round(100 * n_agend / total, 1)
tx_vacuo  = round(100 * n_vacuo / total, 1)
tx_engaj  = round(100 * n_engaj / total, 1)

funil     = df_conv["Estágio Funil"].value_counts().sort_index().to_dict()
motivos   = df_conv[df_conv["Agendou"]=="NÃO"]["Motivo Perda"].value_counts().to_dict()
status_c  = df_conv["Status"].value_counts().to_dict()
utm_c     = df_conv["UTM Source"].value_counts().to_dict()
rem_c     = df_msgs["Remetente"].value_counts().to_dict()
vacuo_pt  = df_conv[df_conv["Vacuo"]=="SIM"]["Msgs IA"].apply(
    lambda x: "Abertura (1ª msg)" if x <= 1 else f"Após {x} msgs da IA"
).value_counts().to_dict()

# Por dia
df_conv["dia"] = pd.to_datetime(df_conv["Criado Em"], utc=True, errors="coerce").dt.day
por_dia = df_conv["dia"].value_counts().sort_index().to_dict()

print(f"\nTotal: {total} | Agendou: {n_agend} ({tx_conv}%) | Vácuo: {n_vacuo} ({tx_vacuo}%) | Humano: {n_humano}")
print("Funil:", funil)

# Salva JSON de stats para o dashboard
stats = {
    "clinica": CLINICA, "agente": AGENTE, "cidade": CIDADE, "canal": CANAL_NUM,
    "total": int(total), "agendou": int(n_agend), "nao_agendou": int(total - n_agend),
    "vacuo": int(n_vacuo), "respondeu": int(n_engaj), "humano": int(n_humano),
    "tx_conv": float(tx_conv), "tx_vacuo": float(tx_vacuo),
    "funil": {k: int(v) for k,v in funil.items()},
    "motivos": {k: int(v) for k,v in list(motivos.items())[:7]},
    "status": {k: int(v) for k,v in status_c.items()},
    "utm": {k: int(v) for k,v in utm_c.items()},
    "rem_counts": {k: int(v) for k,v in rem_c.items()},
    "por_dia": {str(k): int(v) for k,v in por_dia.items()},
    "vacuo_ponto": {k: int(v) for k,v in vacuo_pt.items()},
    "total_msgs": int(len(df_msgs)),
}
with open(OUT_JSON, "w", encoding="utf-8") as f:
    json.dump(stats, f, ensure_ascii=False, indent=2)
print(f"Stats JSON: {OUT_JSON}")

# ── EXCEL ────────────────────────────────────────────────────────────────────
print("\nGerando Excel...")
cores = df_conv.pop("_cor").tolist()
C = {
    "h_dark": "1A252F", "h_mid": "2C3E50", "h_blue": "2E86C1",
    "agendou": "A9DFBF", "perdeu": "FADBD8", "parcial": "FDEBD0",
}

with pd.ExcelWriter(OUT_EXCEL, engine="openpyxl") as writer:

    # Aba 1: Resumo
    ws1 = writer.book.create_sheet("Resumo Executivo")

    def hdr(ws, row, text, cols=3, fg=C["h_mid"]):
        c = ws.cell(row, 1, text)
        c.fill = PatternFill("solid", fgColor=fg)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

    def dado(ws, row, label, valor, cor=None, bold=False):
        lc = ws.cell(row, 1, label); lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(vertical="center", wrap_text=True)
        vc = ws.cell(row, 2, valor); vc.font = Font(bold=bold, size=10)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        if cor: vc.fill = PatternFill("solid", fgColor=cor)

    def texto_row(ws, row, label, txt, cor=None):
        lc = ws.cell(row, 1, label); lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(vertical="top")
        tc = ws.cell(row, 2, txt); tc.font = Font(size=10)
        tc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 50
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        if cor:
            for col in (1,2): ws.cell(row,col).fill = PatternFill("solid",fgColor=cor)

    r = 1
    hdr(ws1, r, f"RELATÓRIO DE IA — {CLINICA} | CANAL {CANAL_NUM} | JUNHO 2026", 3, fg=C["h_dark"]); r+=1
    hdr(ws1, r, f"Agente: {AGENTE} (IA) | {CIDADE} | 01 a 19 de junho de 2026", 3, fg=C["h_blue"]); r+=2

    hdr(ws1, r, "VISÃO GERAL", 3); r+=1
    dado(ws1, r, "Total de leads recebidos", str(total)); r+=1
    dado(ws1, r, "Leads que responderam de verdade", f"{n_engaj}  ({tx_engaj}%)"); r+=1
    dado(ws1, r, "Agendamentos realizados", f"{n_agend}  ({tx_conv}%)", cor=C["agendou"], bold=True); r+=1
    dado(ws1, r, "Perdas", f"{total - n_agend}  ({round(100*(total-n_agend)/total,1)}%)", cor=C["perdeu"]); r+=1
    dado(ws1, r, "Leads no vácuo (nunca responderam)", f"{n_vacuo}  ({tx_vacuo}%)", cor=C["perdeu"]); r+=1
    dado(ws1, r, "Intervenções humanas", f"{n_humano}" + (" — ZERO" if n_humano == 0 else "")); r+=2

    hdr(ws1, r, "FUNIL DE CONVERSÃO", 3); r+=1
    for est_k, qtd in sorted(funil.items()):
        pct = round(100*qtd/total, 1)
        dado(ws1, r, f"  {est_k}", f"{qtd}  ({pct}%)"); r+=1
    r+=1

    hdr(ws1, r, "PRINCIPAIS MOTIVOS DE PERDA", 3); r+=1
    for mot_k, qtd in list(motivos.items())[:6]:
        dado(ws1, r, f"  • {mot_k[:65]}", str(qtd)); r+=1
    r+=1

    hdr(ws1, r, "DIAGNÓSTICO", 3, fg=C["h_dark"]); r+=1
    diags = [
        ("Taxa de conversão",
         f"{tx_conv}% — {'CRÍTICO' if n_agend/total < 0.1 else 'Abaixo do esperado' if n_agend/total < 0.15 else 'Dentro do esperado'}. "
         f"Um SDR IA bem ajustado deve converter entre 15–25% dos leads qualificados.",
         "FADBD8" if n_agend/total < 0.1 else "FFF9C4"),
        ("Leads no vácuo",
         f"{n_vacuo} leads ({tx_vacuo}%) nunca responderam depois da primeira mensagem da IA. "
         f"Indica desalinhamento entre o criativo do anúncio e a abordagem da IA na abertura.",
         "FADBD8" if tx_vacuo > 30 else None),
        ("Maior gargalo do funil",
         f"Estágio com maior abandono: {max(funil, key=funil.get)} ({funil[max(funil, key=funil.get)]} leads). "
         f"Revisar a transição desse estágio para o próximo.",
         "FDEBD0"),
        ("Intervenção humana",
         f"{n_humano} conversa(s) com humano em {total} total. "
         + ("Sem escalonamento — IA tratou tudo sozinha." if n_humano == 0
            else f"Humano interveio em {round(100*n_humano/total,1)}% dos casos."),
         "FFF9C4"),
    ]
    for label, txt, cor in diags:
        texto_row(ws1, r, label, txt, cor=cor); r+=1

    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 70
    ws1.column_dimensions["C"].width = 20

    # Aba 2: Conversas
    df_conv.to_excel(writer, sheet_name="Conversas", index=False)
    ws2 = writer.sheets["Conversas"]
    hf = PatternFill("solid", fgColor=C["h_mid"])
    hft = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws2[1]:
        cell.fill = hf; cell.font = hft
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_agend  = df_conv.columns.get_loc("Agendou") + 1
    col_funil  = df_conv.columns.get_loc("Estágio Funil") + 1
    col_vacuo  = df_conv.columns.get_loc("Vacuo") + 1
    funil_cor  = {"E5":"A9DFBF","E4":"FDEBD0","E3":"FDEBD0","E2":"FFF9C4","E1":"D6EAF8","E0":"FADBD8"}
    for i, cor in enumerate(cores, start=2):
        ws2.row_dimensions[i].height = 32
        ws2.cell(i, col_agend).fill = PatternFill("solid",
            fgColor="A9DFBF" if ws2.cell(i,col_agend).value=="SIM" else "FADBD8")
        ws2.cell(i, col_vacuo).fill = PatternFill("solid",
            fgColor="FADBD8" if ws2.cell(i,col_vacuo).value=="SIM" else "EAFAF1")
        ev = ws2.cell(i, col_funil).value or ""
        fc = next((v for k,v in funil_cor.items() if k in ev), None)
        if fc: ws2.cell(i, col_funil).fill = PatternFill("solid", fgColor=fc)
        ws2.cell(i, df_conv.columns.get_loc("Resumo Paciente")+1).alignment = Alignment(wrap_text=True, vertical="top")
    larguras = {
        "ID Conversa":38,"Contato":24,"Telefone":18,"Status":14,"Criado Em":22,
        "Encerrado Em":22,"UTM Source":14,"UTM Campaign":30,"Estágio Funil":26,
        "Agendou":10,"Vacuo":10,"Motivo Perda":52,"Msgs Paciente":13,"Msgs IA":10,
        "Msgs Sistema":12,"Msgs Humano":12,"Houve Humano":13,
        "Resumo Paciente":65,"Última Msg IA":60,
    }
    for ci, col in enumerate(df_conv.columns, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = larguras.get(col, 15)
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    # Aba 3: Mensagens
    df_msgs.to_excel(writer, sheet_name="Mensagens", index=False)
    ws3 = writer.sheets["Mensagens"]
    for cell in ws3[1]:
        cell.fill = hf; cell.font = hft
        cell.alignment = Alignment(horizontal="center", vertical="center")
    col_rem = df_msgs.columns.get_loc("Remetente") + 1
    col_txt = df_msgs.columns.get_loc("Texto") + 1
    rem_cor = {"PACIENTE":"EAFAF1","IA":"EBF5FB","SISTEMA":"F2F3F4","HUMANO":"FDF2F8"}
    for i, row in enumerate(df_msgs.itertuples(), start=2):
        fc = rem_cor.get(row.Remetente)
        if fc: ws3.cell(i, col_rem).fill = PatternFill("solid", fgColor=fc)
        ws3.cell(i, col_txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[i].height = 28
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 70
    ws3.column_dimensions["G"].width = 22
    ws3.freeze_panes = "A2"

    wb = writer.book
    wb.move_sheet("Resumo Executivo", offset=-len(wb.sheetnames))

# Remove checkpoint ao concluir com sucesso
if os.path.exists(OUT_CKPT):
    os.remove(OUT_CKPT)
    print("Checkpoint removido.")

print(f"\nExcel: {OUT_EXCEL}")
print(f"Total: {total} sessoes | {len(df_msgs)} mensagens")
print(f"Agendamentos: {n_agend} ({tx_conv}%) | Vacuo: {n_vacuo} ({tx_vacuo}%) | Humano: {n_humano}")
